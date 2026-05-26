import sqlite3
import contextvars
from enum import StrEnum

_current_conn = contextvars.ContextVar("_current_conn", default=None)

import os
import re
import json
import datetime
import logging
import contextvars
from collections import defaultdict
from typing import Optional

# ── Current user context ───────────────────────────────────────────────────────
# Set by the identity middleware in main.py before each request.
# ContextVar is used instead of a plain global so that concurrent async
# requests each carry their own value — a module-level str would be shared
# across coroutines and could be overwritten by a second request between the
# middleware set and the route handler's DB call.
_current_user: contextvars.ContextVar[str] = contextvars.ContextVar(
    'current_user', default='system'
)

def set_current_user(name: str) -> None:
    _current_user.set(name or 'unknown')

def get_current_user() -> str:
    return _current_user.get()

# ── Audit log helper ───────────────────────────────────────────────────────────
def log_change(conn, entity_type: str, entity_id: str, action: str,
               field: str = None, old_val=None, new_val=None) -> None:
    """
    Write one row to change_log inside the caller's transaction.
    Always called within an existing `with get_conn() as conn:` block so
    if the write fails and rolls back, this entry rolls back too.
    """
    try:
        conn.execute(
            """INSERT INTO change_log (ts, user, entity_type, entity_id, action, field, old_val, new_val)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                datetime.datetime.now().isoformat(timespec='seconds'),
                _current_user.get(),
                entity_type,
                str(entity_id),
                action,
                field,
                str(old_val) if old_val is not None else None,
                str(new_val) if new_val is not None else None,
            )
        )
    except Exception as e:
        logging.warning("log_change failed (%s %s): %s", entity_type, entity_id, e)

def _norm_for_diff(val) -> str:
    """Canonical string for change-detection comparison.
    Treats None, '', and any numeric zero (0, 0.0, '0', '0.0') as equivalent
    so that a first-save of default zeros against a NULL column — or a
    string '0' from a form against an integer 0 from the DB — never produces
    spurious audit entries.
    """
    if val is None:
        return ''
    s = str(val).strip()
    if s in ('', 'None'):
        return ''
    try:
        f = float(s)
        return '' if f == 0.0 else f'{f:g}'   # ':g' strips trailing zeros
    except (ValueError, TypeError):
        return s

def _diff_log(conn, entity_type: str, entity_id: str, old: dict, new: dict,
              fields: list) -> None:
    """Log one change_log row for each field in `fields` that actually changed.
    Fields absent from `new` are skipped — a missing key is not treated as
    a deletion, so partial saves never produce spurious 'X → None' entries.
    """
    for f in fields:
        if f not in new:          # field not being updated in this save — skip
            continue
        ov = _norm_for_diff(old.get(f))
        nv = _norm_for_diff(new.get(f))
        if ov != nv:
            # Pass the raw values to log_change so the UI shows real before/after
            log_change(conn, entity_type, entity_id, 'update', f,
                       old.get(f), new.get(f))


# ── Audit log queries ──────────────────────────────────────────────────────────
def get_change_log(entity_type: str = None, entity_id: str = None,
                   user: str = None, date_from: str = None,
                   date_to: str = None, limit: int = 200,
                   offset: int = 0) -> list[dict]:
    q      = "SELECT * FROM change_log WHERE 1=1"
    params = []
    if entity_type: q += " AND entity_type=?";  params.append(entity_type)
    if entity_id:   q += " AND entity_id LIKE ?"; params.append(f'%{entity_id}%')
    if user:        q += " AND user=?";           params.append(user)
    if date_from:   q += " AND ts >= ?";          params.append(date_from)
    if date_to:     q += " AND ts <= ?";          params.append(date_to + 'T23:59:59')
    q += " ORDER BY id DESC LIMIT ? OFFSET ?"
    params += [limit, offset]
    with get_conn() as conn:
        return [dict(r) for r in conn.execute(q, params).fetchall()]

def get_change_log_users() -> list[str]:
    with get_conn() as conn:
        return [r[0] for r in conn.execute(
            "SELECT DISTINCT user FROM change_log ORDER BY user"
        ).fetchall()]

DB_PATH = os.path.join(os.path.dirname(__file__), "lems_core.db")

class PartType(StrEnum):
    PRT = "PRT"
    RAW = "RAW"
    FAB = "FAB"
    ASSY = "ASSY"

class PartCategory(StrEnum):
    DLT = "DLT"
    SBX = "SBX"
    GRV = "GRV"
    CAL = "CAL"
    BLC = "BLC"
    LAB = "LAB"
    STD = "STD"

class PartStatus(StrEnum):
    ACTIVE = "ACTIVE"
    OBSOLETE = "OBSOLETE"

class ItemType(StrEnum):
    STANDARD = "STANDARD"
    ADDITIONAL = "ADDITIONAL"
    OPTIONAL = "OPTIONAL"

TYPES      = list(PartType)
CATEGORIES = list(PartCategory)
PART_STATUSES = list(PartStatus)
ITEM_TYPES = list(ItemType)
PART_TYPES_WITH_BOM    = {PartType.FAB, PartType.ASSY}
PART_TYPES_STATIC_COST = {PartType.PRT, PartType.RAW}

def get_conn():
    conn = _current_conn.get()
    if conn is not None:
        return conn
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def init_db():
    with get_conn() as conn:
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA synchronous=NORMAL;")
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS parts (
            part_id      TEXT PRIMARY KEY,
            type         TEXT NOT NULL,
            category     TEXT NOT NULL,
            base_desc    TEXT NOT NULL,
            size_spec    TEXT DEFAULT '',
            variant      TEXT DEFAULT '',
            plain_desc   TEXT DEFAULT '',
            supplier     TEXT DEFAULT '',
            brand_mfg    TEXT DEFAULT '',
            supplier_pn  TEXT DEFAULT '',
            uom          TEXT DEFAULT 'ea',
            pkg_size     REAL DEFAULT 1,
            pkg_cost     REAL DEFAULT 0,
            unit_cost    REAL DEFAULT 0,
            labor_hrs    REAL DEFAULT 0,
            qty_on_hand  REAL DEFAULT 0,
            cost         REAL DEFAULT 0,
            on_hand      REAL DEFAULT 0,
            status       TEXT DEFAULT 'ACTIVE',
            rolled_labor_hrs REAL DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS bom (
            parent_id  TEXT NOT NULL REFERENCES parts(part_id) ON UPDATE CASCADE,
            child_id   TEXT NOT NULL REFERENCES parts(part_id) ON UPDATE CASCADE,
            qty        REAL NOT NULL DEFAULT 1,
            sort_order INTEGER DEFAULT 0,
            PRIMARY KEY (parent_id, child_id)
        );

        CREATE TABLE IF NOT EXISTS projects (
            project_id TEXT PRIMARY KEY,
            status     TEXT DEFAULT 'ACTIVE',
            customer   TEXT DEFAULT '',
            notes      TEXT DEFAULT '',
            labor_rate REAL DEFAULT 25,
            markup     REAL DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS project_items (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id TEXT NOT NULL REFERENCES projects(project_id) ON DELETE CASCADE,
            part_id    TEXT NOT NULL REFERENCES parts(part_id),
            qty        REAL DEFAULT 1,
            picked     INTEGER DEFAULT 0,
            item_type  TEXT DEFAULT 'ADDITIONAL',
            box_num    TEXT DEFAULT '',
            discount_pct  REAL DEFAULT 0,
            discount_flat REAL DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS project_boxes (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id TEXT NOT NULL REFERENCES projects(project_id) ON DELETE CASCADE,
            box_num    TEXT NOT NULL,
            weight     REAL DEFAULT 0,
            pallet_num TEXT DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS project_pallets (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id TEXT NOT NULL REFERENCES projects(project_id) ON DELETE CASCADE,
            pallet_num TEXT NOT NULL,
            weight     REAL DEFAULT 0,
            dimensions TEXT DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS project_other_items (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id     TEXT NOT NULL REFERENCES projects(project_id) ON DELETE CASCADE,
            description    TEXT DEFAULT '',
            cost           REAL DEFAULT 0,
            labor_hrs      REAL DEFAULT 0,
            apply_markup   INTEGER DEFAULT 0,
            box_num        TEXT DEFAULT '',
            discount_pct   REAL DEFAULT 0,
            discount_flat  REAL DEFAULT 0,
            show_on_proforma INTEGER DEFAULT 0,
            sort_order     INTEGER DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS project_pick_status (
            project_id TEXT NOT NULL REFERENCES projects(project_id) ON DELETE CASCADE,
            part_id    TEXT NOT NULL,
            picked     INTEGER DEFAULT 0,
            PRIMARY KEY (project_id, part_id)
        );

        CREATE TABLE IF NOT EXISTS project_quotes (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id       TEXT NOT NULL UNIQUE REFERENCES projects(project_id) ON DELETE CASCADE,
            version          INTEGER DEFAULT 1,
            status           TEXT DEFAULT 'DRAFT',
            currency         TEXT DEFAULT 'USD',
            -- Cost modifiers
            overhead_rate    REAL DEFAULT 1.0,
            markup_pct       REAL DEFAULT 0,
            labor_rate_quoted REAL DEFAULT 0,
            -- Additional line items
            freight_inbound  REAL DEFAULT 0,
            freight_outbound REAL DEFAULT 0,
            cal_gases_cost   REAL DEFAULT 0,
            cal_gases_freight REAL DEFAULT 0,
            training_days    REAL DEFAULT 0,
            training_cost    REAL DEFAULT 0,
            training_notes   TEXT DEFAULT '',
            -- Frozen BOM costs snapshot
            frozen_material  REAL DEFAULT 0,
            frozen_labor_hrs REAL DEFAULT 0,
            costs_frozen     INTEGER DEFAULT 0,
            frozen_at        TEXT,
            -- Custom line items JSON: [{desc, cost, show_on_proforma}]
            other_items      TEXT DEFAULT '[]',
            -- Notes / text
            internal_notes   TEXT DEFAULT '',
            proforma_header  TEXT DEFAULT '',
            proforma_footer  TEXT DEFAULT '',
            quoted_total     REAL DEFAULT 0,
            gross_margin_pct REAL DEFAULT 0,
            total_internal   REAL DEFAULT 0,
            created_at       TEXT DEFAULT (datetime('now')),
            updated_at       TEXT DEFAULT (datetime('now'))
        );

        CREATE INDEX IF NOT EXISTS idx_bom_child ON bom(child_id);
        CREATE INDEX IF NOT EXISTS idx_bom_parent ON bom(parent_id);
        CREATE INDEX IF NOT EXISTS idx_project_items_pid ON project_items(project_id);
        CREATE INDEX IF NOT EXISTS idx_project_pick_pid ON project_pick_status(project_id);
        
        CREATE TABLE IF NOT EXISTS part_cost_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            part_id TEXT NOT NULL REFERENCES parts(part_id) ON DELETE CASCADE,
            old_cost REAL NOT NULL,
            new_cost REAL NOT NULL,
            changed_at TEXT DEFAULT (datetime('now'))
        );
        
        CREATE TABLE IF NOT EXISTS part_attachments (
            id TEXT PRIMARY KEY,
            part_id TEXT NOT NULL REFERENCES parts(part_id) ON DELETE CASCADE,
            filename TEXT NOT NULL,
            original_filename TEXT NOT NULL,
            mime_type TEXT NOT NULL,
            size_bytes INTEGER NOT NULL,
            uploaded_by TEXT NOT NULL,
            uploaded_at TEXT DEFAULT (datetime('now'))
        );
        
        CREATE TABLE IF NOT EXISTS change_log (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            ts          TEXT NOT NULL,
            user        TEXT NOT NULL,
            entity_type TEXT NOT NULL,
            entity_id   TEXT NOT NULL,
            action      TEXT NOT NULL,
            field       TEXT,
            old_val     TEXT,
            new_val     TEXT
        );
        """)
        # Migrate existing DBs that may be missing newer columns
        _run_migrations(conn)

def _run_migrations(conn):
    conn.execute("CREATE TABLE IF NOT EXISTS schema_migrations (version INTEGER PRIMARY KEY)")
    row = conn.execute("SELECT MAX(version) FROM schema_migrations").fetchone()
    current_version = row[0] if row[0] else 0

    migrations = [
        (1, _migration_v1_legacy),
        (2, _migration_v2_other_items),
        (3, _migration_v3_on_update_cascade),
        (4, _migration_v4_data_integrity),
        (5, _migration_v5_cost_history),
        (6, _migration_v6_attachments),
    ]

    for version, func in migrations:
        if version > current_version:
            try:
                func(conn)
                conn.execute("INSERT INTO schema_migrations (version) VALUES (?)", (version,))
                conn.commit()
            except Exception as e:
                conn.rollback()
                print(f"Migration v{version} failed: {e}")
                raise

def _migration_v1_legacy(conn):
    def col_exists(table, col):
        cols = [r[1] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()]
        return col in cols

    stmts = [
        ("parts", "status",       "ALTER TABLE parts ADD COLUMN status TEXT DEFAULT 'ACTIVE'"),
        ("parts", "rolled_labor_hrs", "ALTER TABLE parts ADD COLUMN rolled_labor_hrs REAL DEFAULT 0"),
        ("bom",   "sort_order",   "ALTER TABLE bom ADD COLUMN sort_order INTEGER DEFAULT 0"),
        ("projects", "markup",    "ALTER TABLE projects ADD COLUMN markup REAL DEFAULT 0"),
        ("projects", "created_at","ALTER TABLE projects ADD COLUMN created_at TEXT DEFAULT (datetime('now'))"),
        ("parts", "supplier_2",   "ALTER TABLE parts ADD COLUMN supplier_2 TEXT DEFAULT ''"),
        ("parts", "brand_mfg_2",  "ALTER TABLE parts ADD COLUMN brand_mfg_2 TEXT DEFAULT ''"),
        ("parts", "supplier_pn_2","ALTER TABLE parts ADD COLUMN supplier_pn_2 TEXT DEFAULT ''"),
        ("parts", "pkg_size_2",   "ALTER TABLE parts ADD COLUMN pkg_size_2 REAL DEFAULT 1"),
        ("parts", "pkg_cost_2",   "ALTER TABLE parts ADD COLUMN pkg_cost_2 REAL DEFAULT 0"),
        ("parts", "unit_cost_2",  "ALTER TABLE parts ADD COLUMN unit_cost_2 REAL DEFAULT 0"),
        ("parts", "use_alt_supplier", "ALTER TABLE parts ADD COLUMN use_alt_supplier INTEGER DEFAULT 0"),
        ("parts", "last_cost_date",   "ALTER TABLE parts ADD COLUMN last_cost_date TEXT DEFAULT ''"),
        ("project_quotes", "markup_pct", "ALTER TABLE project_quotes ADD COLUMN markup_pct REAL DEFAULT 0"),
        ("project_items", "item_type", "ALTER TABLE project_items ADD COLUMN item_type TEXT DEFAULT 'ADDITIONAL'"),
        ("project_items", "box_num", "ALTER TABLE project_items ADD COLUMN box_num TEXT DEFAULT ''"),
        ("project_quotes", "discount_pct",  "ALTER TABLE project_quotes ADD COLUMN discount_pct  REAL DEFAULT 0"),
        ("project_quotes", "discount_flat", "ALTER TABLE project_quotes ADD COLUMN discount_flat REAL DEFAULT 0"),
        ("project_quotes", "discount_note", "ALTER TABLE project_quotes ADD COLUMN discount_note TEXT DEFAULT ''"),
        ("project_quotes", "labor_rate_quoted", "ALTER TABLE project_quotes ADD COLUMN labor_rate_quoted REAL DEFAULT 0"),
        ("project_quotes", "total_internal",   "ALTER TABLE project_quotes ADD COLUMN total_internal REAL DEFAULT 0"),
        ("project_quotes", "quoted_total",     "ALTER TABLE project_quotes ADD COLUMN quoted_total REAL DEFAULT 0"),
        ("project_quotes", "gross_margin_pct", "ALTER TABLE project_quotes ADD COLUMN gross_margin_pct REAL DEFAULT 0"),
        ("project_items", "discount_pct",  "ALTER TABLE project_items ADD COLUMN discount_pct REAL DEFAULT 0"),
        ("project_items", "discount_flat", "ALTER TABLE project_items ADD COLUMN discount_flat REAL DEFAULT 0"),
        ("parts",               "qty_on_order", "ALTER TABLE parts ADD COLUMN qty_on_order REAL DEFAULT 0"),
        ("parts",               "order_eta",    "ALTER TABLE parts ADD COLUMN order_eta TEXT DEFAULT ''"),
        ("project_pick_status", "picked_qty",   "ALTER TABLE project_pick_status ADD COLUMN picked_qty REAL DEFAULT 0"),
        ("change_log", "id", """CREATE TABLE IF NOT EXISTS change_log (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            ts          TEXT NOT NULL,
            user        TEXT NOT NULL,
            entity_type TEXT NOT NULL,
            entity_id   TEXT NOT NULL,
            action      TEXT NOT NULL,
            field       TEXT,
            old_val     TEXT,
            new_val     TEXT
        )"""),
    ]
    for table, col, sql in stmts:
        try:
            if sql.strip().upper().startswith('CREATE TABLE'):
                exists = conn.execute("SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name=?",(table,)).fetchone()[0]
                if not exists: conn.executescript(sql)
            elif not col_exists(table, col):
                conn.execute(sql)
        except Exception:
            pass

    conn.execute("UPDATE bom SET sort_order = rowid WHERE sort_order = 0 AND parent_id != '' AND parent_id IS NOT NULL")
    try:
        conn.execute("UPDATE project_items SET item_type = 'STANDARD' WHERE item_type IS NULL OR item_type = 'ADDITIONAL' AND part_id LIKE '%-STDPKG%'")
    except Exception:
        pass

def _migration_v2_other_items(conn):
    import json
    try:
        quotes = conn.execute("SELECT project_id, other_items FROM project_quotes WHERE other_items IS NOT NULL AND other_items != '[]'").fetchall()
        for q in quotes:
            existing = conn.execute("SELECT COUNT(*) FROM project_other_items WHERE project_id=?", (q['project_id'],)).fetchone()[0]
            if existing > 0: continue
            items = json.loads(q['other_items'] or '[]')
            for idx, i in enumerate(items):
                if not isinstance(i, dict): continue
                conn.execute(
                    """INSERT INTO project_other_items
                       (project_id, description, cost, labor_hrs, apply_markup, box_num, discount_pct, discount_flat, show_on_proforma, sort_order)
                       VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    (q['project_id'], (i.get('desc') or i.get('description') or '').strip(), float(i.get('cost') or 0), float(i.get('labor_hrs') or 0), 1 if i.get('apply_markup') else 0, (i.get('box_num') or '').strip(), float(i.get('discount_pct') or 0), float(i.get('discount_flat') or 0), 1 if i.get('show_on_proforma') else 0, idx)
                )
    except Exception:
        pass

def _migration_v4_data_integrity(conn):
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS quote_snapshots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id TEXT NOT NULL REFERENCES projects(project_id) ON DELETE CASCADE,
            version INTEGER NOT NULL,
            snapshot_json TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now')),
            UNIQUE(project_id, version)
        );
    """)
    try:
        conn.execute("ALTER TABLE parts ADD COLUMN deleted_at TEXT DEFAULT NULL")
    except Exception:
        pass

def _migration_v5_cost_history(conn):
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS part_cost_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            part_id TEXT NOT NULL REFERENCES parts(part_id) ON DELETE CASCADE ON UPDATE CASCADE,
            old_cost REAL NOT NULL,
            new_cost REAL NOT NULL,
            changed_at TEXT DEFAULT (datetime('now'))
        );
        CREATE INDEX IF NOT EXISTS idx_part_cost_history_part_id ON part_cost_history(part_id);
    """)

def _migration_v6_attachments(conn):
    conn.execute('''
        CREATE TABLE IF NOT EXISTS part_attachments (
            id TEXT PRIMARY KEY,
            part_id TEXT,
            filename TEXT,
            original_filename TEXT,
            mime_type TEXT,
            size_bytes INTEGER,
            uploaded_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            uploaded_by TEXT,
            FOREIGN KEY(part_id) REFERENCES parts(part_id) ON UPDATE CASCADE ON DELETE CASCADE
        )
    ''')


def _migration_v3_on_update_cascade(conn):
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS new_project_items (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id TEXT NOT NULL REFERENCES projects(project_id) ON DELETE CASCADE,
            part_id    TEXT NOT NULL REFERENCES parts(part_id) ON UPDATE CASCADE,
            qty        REAL DEFAULT 1,
            picked     INTEGER DEFAULT 0,
            item_type  TEXT DEFAULT 'ADDITIONAL',
            box_num    TEXT DEFAULT '',
            discount_pct  REAL DEFAULT 0,
            discount_flat REAL DEFAULT 0
        );
        INSERT INTO new_project_items (id, project_id, part_id, qty, picked, item_type, box_num, discount_pct, discount_flat)
        SELECT id, project_id, part_id, qty, picked, item_type, box_num, discount_pct, discount_flat FROM project_items;
        DROP TABLE project_items;
        ALTER TABLE new_project_items RENAME TO project_items;
        CREATE INDEX IF NOT EXISTS idx_project_items_pid ON project_items(project_id);

        CREATE TABLE IF NOT EXISTS new_project_pick_status (
            project_id TEXT NOT NULL REFERENCES projects(project_id) ON DELETE CASCADE,
            part_id    TEXT NOT NULL REFERENCES parts(part_id) ON UPDATE CASCADE,
            picked     INTEGER DEFAULT 0,
            picked_qty REAL DEFAULT 0,
            PRIMARY KEY (project_id, part_id)
        );
        INSERT INTO new_project_pick_status (project_id, part_id, picked, picked_qty)
        SELECT project_id, part_id, picked, picked_qty FROM project_pick_status;
        DROP TABLE project_pick_status;
        ALTER TABLE new_project_pick_status RENAME TO project_pick_status;
        CREATE INDEX IF NOT EXISTS idx_project_pick_pid ON project_pick_status(project_id);
    """)

# ── Part ID helpers ────────────────────────────────────────────────────────────

def clean(text: str) -> str:
    return re.sub(r'[^A-Z0-9\-\/\.]', '', str(text).upper().strip())

def build_part_id(type_: str, category: str, base_desc: str,
                  size_spec: str = '', variant: str = '') -> str:
    segments = [clean(type_), clean(category), clean(base_desc)]
    if size_spec: segments.append(clean(size_spec))
    if variant:   segments.append(clean(variant))
    return '-'.join(s for s in segments if s)

# ── Parts CRUD ─────────────────────────────────────────────────────────────────

def get_all_parts(search: str = '', type_filter: str = '',
                  cat_filter: str = '', include_obsolete: bool = True) -> list[dict]:
    with get_conn() as conn:
        q, params = "SELECT * FROM parts WHERE 1=1", []
        if search:
            q += " AND (part_id LIKE ? OR plain_desc LIKE ? OR supplier LIKE ? OR brand_mfg LIKE ?)"
            s = f'%{search}%'
            params += [s, s, s, s]
        if type_filter:
            q += " AND type = ?"; params.append(type_filter)
        if cat_filter:
            q += " AND category = ?"; params.append(cat_filter)
        if not include_obsolete:
            q += " AND (status IS NULL OR status != 'OBSOLETE')"
        q += " ORDER BY type, category, base_desc"
        return [dict(r) for r in conn.execute(q, params).fetchall()]

def get_part(part_id: str) -> Optional[dict]:
    with get_conn() as conn:
        return _get_part(conn, part_id)

def _get_part(conn, part_id: str) -> Optional[dict]:
    row = conn.execute("SELECT * FROM parts WHERE part_id = ?", (part_id,)).fetchone()
    return dict(row) if row else None

def upsert_part(data: dict, orig_part_id: str = None) -> tuple[bool, str]:
    """
    Create or update a part.
    If orig_part_id is provided (edit mode), always save to that ID —
    the ID-segment fields (type/category/base_desc etc.) are informational only
    and do not change the stored part_id on edit.
    """
    if orig_part_id:
        # Edit mode: use the original ID, never rename
        part_id = orig_part_id
    else:
        part_id = build_part_id(
            data.get('type',''), data.get('category',''), data.get('base_desc',''),
            data.get('size_spec',''), data.get('variant','')
        )
    if not part_id or len(part_id) < 5:
        return False, "Invalid part ID — check type, category, and base description."

    pkg_size  = float(data.get('pkg_size')  or 1)
    pkg_cost  = float(data.get('pkg_cost')  or 0)
    unit_cost = float(data.get('unit_cost') or 0)

    pkg_size_2  = float(data.get('pkg_size_2')  or 1)
    pkg_cost_2  = float(data.get('pkg_cost_2')  or 0)
    unit_cost_2 = float(data.get('unit_cost_2') or 0)
    use_alt     = int(bool(data.get('use_alt_supplier') and data.get('use_alt_supplier') not in ('0', '', False)))

    if data.get('type') in PART_TYPES_STATIC_COST:
        if pkg_size > 0 and pkg_cost > 0:
            unit_cost = round(pkg_cost / pkg_size, 6)
        if pkg_size_2 > 0 and pkg_cost_2 > 0:
            unit_cost_2 = round(pkg_cost_2 / pkg_size_2, 6)

    # If alt supplier is active, use alt unit cost as the effective unit_cost
    if use_alt and unit_cost_2 > 0:
        unit_cost = unit_cost_2

    # Auto-stamp last_cost_date when costs are present
    last_cost_date = data.get('last_cost_date', '').strip()
    if not last_cost_date and (pkg_cost > 0 or pkg_cost_2 > 0):
        last_cost_date = datetime.date.today().isoformat()

    LOGGED_FIELDS = [
        'type','category','base_desc','size_spec','variant','plain_desc',
        'supplier','brand_mfg','supplier_pn','uom','pkg_size','pkg_cost',
        'unit_cost','labor_hrs','qty_on_hand','status',
        'supplier_2','brand_mfg_2','supplier_pn_2',
        'pkg_size_2','pkg_cost_2','unit_cost_2','use_alt_supplier','last_cost_date',
    ]

    with get_conn() as conn:
        existing = conn.execute(
            "SELECT * FROM parts WHERE part_id = ?", (part_id,)
        ).fetchone()
        old_row = dict(existing) if existing else None

        # Refresh last_cost_date if primary or alt cost actually changed
        if old_row:
            old_primary = float(old_row.get('pkg_cost') or 0)
            old_alt     = float(old_row.get('pkg_cost_2') or 0)
            if (pkg_cost != old_primary or pkg_cost_2 != old_alt) and (pkg_cost > 0 or pkg_cost_2 > 0):
                last_cost_date = datetime.date.today().isoformat()

        fields = (
            data.get('type'), data.get('category'), data.get('base_desc'),
            data.get('size_spec',''), data.get('variant',''), data.get('plain_desc',''),
            data.get('supplier',''), data.get('brand_mfg',''), data.get('supplier_pn',''),
            data.get('uom','ea'), pkg_size, pkg_cost, unit_cost,
            float(data.get('labor_hrs') or 0), float(data.get('qty_on_hand') or 0),
            data.get('status','ACTIVE'),
            data.get('supplier_2',''), data.get('brand_mfg_2',''), data.get('supplier_pn_2',''),
            pkg_size_2, pkg_cost_2, unit_cost_2, use_alt, last_cost_date,
        )
        if old_row:
            if not orig_part_id:
                return False, f"Part ID '{part_id}' already exists. Use different type, category, description, size, or variant."
            conn.execute("""UPDATE parts SET
                type=?,category=?,base_desc=?,size_spec=?,variant=?,plain_desc=?,
                supplier=?,brand_mfg=?,supplier_pn=?,uom=?,pkg_size=?,pkg_cost=?,
                unit_cost=?,labor_hrs=?,qty_on_hand=?,status=?,
                supplier_2=?,brand_mfg_2=?,supplier_pn_2=?,
                pkg_size_2=?,pkg_cost_2=?,unit_cost_2=?,use_alt_supplier=?,last_cost_date=?
                WHERE part_id=?
            """, (*fields, part_id))
            new_row = dict(zip(LOGGED_FIELDS, fields))
            _diff_log(conn, 'part', part_id, old_row, new_row, LOGGED_FIELDS)
            
            old_uc = float(old_row.get('unit_cost') or 0)
            if unit_cost != old_uc and old_uc > 0:
                conn.execute("INSERT INTO part_cost_history (part_id, old_cost, new_cost) VALUES (?, ?, ?)", (part_id, old_uc, unit_cost))
        else:
            conn.execute("""INSERT INTO parts
                (type,category,base_desc,size_spec,variant,plain_desc,supplier,brand_mfg,
                 supplier_pn,uom,pkg_size,pkg_cost,unit_cost,labor_hrs,qty_on_hand,status,
                 supplier_2,brand_mfg_2,supplier_pn_2,pkg_size_2,pkg_cost_2,unit_cost_2,
                 use_alt_supplier,last_cost_date,part_id)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (*fields, part_id))
            log_change(conn, 'part', part_id, 'create')
    return True, part_id

def delete_part(part_id: str) -> tuple[bool, str]:
    with get_conn() as conn:
        conn.execute("UPDATE parts SET status = 'OBSOLETE' WHERE part_id = ?", (part_id,))
        log_change(conn, 'part', part_id, 'obsolete')
    return True, "Part marked as obsolete."
def update_part_field(part_id: str, field: str, value: str) -> tuple[bool, str]:
    safe = {'plain_desc','supplier','brand_mfg','supplier_pn','uom',
            'pkg_size','pkg_cost','unit_cost','labor_hrs','qty_on_hand','status'}
    if field not in safe:
        return False, f"Field '{field}' is not inline-editable."
    try:
        with get_conn() as conn:
            # Read old value for audit diff before overwriting
            old_row = conn.execute(
                f"SELECT {field} FROM parts WHERE part_id=?", (part_id,)
            ).fetchone()
            old_val = str(old_row[0]) if old_row else None
            conn.execute(f"UPDATE parts SET {field}=? WHERE part_id=?", (value, part_id))
            log_change(conn, 'part', part_id, 'update',
                       field=field, old_val=old_val, new_val=value)
            
            if field == 'unit_cost':
                old_uc = float(old_val or 0)
                new_uc = float(value or 0)
                if new_uc != old_uc and old_uc > 0:
                    conn.execute("INSERT INTO part_cost_history (part_id, old_cost, new_cost) VALUES (?, ?, ?)", (part_id, old_uc, new_uc))
                    
            if field in ('pkg_size', 'pkg_cost'):
                row = conn.execute(
                    "SELECT type,pkg_size,pkg_cost,unit_cost FROM parts WHERE part_id=?", (part_id,)
                ).fetchone()
                if row and row['type'] in PART_TYPES_STATIC_COST:
                    ps, pc = float(row['pkg_size'] or 1), float(row['pkg_cost'] or 0)
                    old_uc = float(row['unit_cost'] or 0)
                    if ps > 0:
                        new_uc = round(pc/ps, 6)
                        conn.execute("UPDATE parts SET unit_cost=? WHERE part_id=?", (new_uc, part_id))
                        if new_uc != old_uc and old_uc > 0:
                            conn.execute("INSERT INTO part_cost_history (part_id, old_cost, new_cost) VALUES (?, ?, ?)", (part_id, old_uc, new_uc))
        return True, "Saved."
    except Exception as e:
        return False, str(e)

def update_part_labor(part_id: str, labor_hrs: float) -> tuple[bool, str]:
    """Update labor_hrs for a part — called from BOM editor."""
    try:
        with get_conn() as conn:
            conn.execute("UPDATE parts SET labor_hrs=? WHERE part_id=?", (labor_hrs, part_id))
        return True, "Labor updated."
    except Exception as e:
        return False, str(e)

# ── BOM CRUD ───────────────────────────────────────────────────────────────────

def get_bom_children(parent_id: str) -> list[dict]:
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT b.child_id, b.qty, b.sort_order,
                   p.plain_desc, p.type, p.unit_cost, p.uom, p.status, p.labor_hrs, p.status
            FROM bom b JOIN parts p ON p.part_id = b.child_id
            WHERE b.parent_id = ?
            ORDER BY b.sort_order, b.child_id
        """, (parent_id,)).fetchall()
        return [dict(r) for r in rows]

def get_bom_parents(child_id: str) -> list[dict]:
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT b.parent_id, p.plain_desc, p.type
            FROM bom b JOIN parts p ON p.part_id = b.parent_id
            WHERE b.child_id = ?
        """, (child_id,)).fetchall()
        return [dict(r) for r in rows]

def add_bom_row(parent_id: str, child_id: str, qty: float) -> tuple[bool, str]:
    if parent_id == child_id:
        return False, "A part cannot be its own child."
    if _would_create_cycle(parent_id, child_id):
        return False, "This would create a circular BOM reference."
    p = get_part(parent_id)
    c = get_part(child_id)
    if not p: return False, f"Parent not found: {parent_id}"
    if not c: return False, f"Child not found: {child_id}"
    if p['type'] not in PART_TYPES_WITH_BOM:
        return False, f"Parent must be FAB or ASSY (got {p['type']})."
    with get_conn() as conn:
        try:
            max_order = conn.execute(
                "SELECT COALESCE(MAX(sort_order),0) FROM bom WHERE parent_id=?", (parent_id,)
            ).fetchone()[0]
            conn.execute(
                "INSERT INTO bom (parent_id,child_id,qty,sort_order) VALUES (?,?,?,?)",
                (parent_id, child_id, qty, max_order + 10)
            )
            log_change(conn, 'bom', parent_id, 'create',
                       field='child_id', old_val=None, new_val=f"{child_id} qty={qty}")
            return True, "Component added."
        except sqlite3.IntegrityError:
            return False, "This relationship already exists."

def update_bom_qty(parent_id: str, child_id: str, qty: float) -> tuple[bool, str]:
    with get_conn() as conn:
        old = conn.execute(
            "SELECT qty FROM bom WHERE parent_id=? AND child_id=?", (parent_id, child_id)
        ).fetchone()
        old_qty = old['qty'] if old else None
        conn.execute("UPDATE bom SET qty=? WHERE parent_id=? AND child_id=?",
                     (qty, parent_id, child_id))
        log_change(conn, 'bom', parent_id, 'update',
                   field=f'qty[{child_id}]', old_val=old_qty, new_val=qty)
    return True, "Quantity updated."

def delete_bom_row(parent_id: str, child_id: str) -> tuple[bool, str]:
    with get_conn() as conn:
        old = conn.execute(
            "SELECT qty FROM bom WHERE parent_id=? AND child_id=?", (parent_id, child_id)
        ).fetchone()
        conn.execute("DELETE FROM bom WHERE parent_id=? AND child_id=?",
                     (parent_id, child_id))
        log_change(conn, 'bom', parent_id, 'delete',
                   field='child_id', old_val=f"{child_id} qty={old['qty'] if old else '?'}", new_val=None)
    return True, "Component removed."

def reorder_bom(parent_id: str, ordered_child_ids: list[str]) -> tuple[bool, str]:
    """Update sort_order for all children of parent_id based on given order."""
    with get_conn() as conn:
        for i, child_id in enumerate(ordered_child_ids):
            conn.execute(
                "UPDATE bom SET sort_order=? WHERE parent_id=? AND child_id=?",
                (i * 10, parent_id, child_id)
            )
    return True, "Order saved."

def _would_create_cycle(parent_id: str, child_id: str) -> bool:
    with get_conn() as conn:
        visited, stack = set(), [child_id]
        while stack:
            node = stack.pop()
            if node == parent_id: return True
            if node in visited: continue
            visited.add(node)
            stack.extend(r[0] for r in conn.execute(
                "SELECT child_id FROM bom WHERE parent_id=?", (node,)
            ).fetchall())
    return False

# ── BOM Tree ───────────────────────────────────────────────────────────────────

def _build_bom_ctx(conn) -> dict:
    parts_rows = conn.execute("SELECT * FROM parts").fetchall()
    parts = {r['part_id']: dict(r) for r in parts_rows}
    bom_rows = conn.execute("SELECT * FROM bom ORDER BY sort_order, child_id").fetchall()
    bom_children = {}
    bom_parents = {}
    for r in bom_rows:
        bom_children.setdefault(r['parent_id'], []).append(dict(r))
        bom_parents.setdefault(r['child_id'], []).append(dict(r))
    return {'parts': parts, 'bom_children': bom_children, 'bom_parents': bom_parents}

def build_bom_tree(parent_id: str, depth: int = 0, visited: set = None, _conn = None, bom_ctx: dict = None) -> dict:
    if bom_ctx is None:
        if _conn is None:
            with get_conn() as conn:
                return build_bom_tree(parent_id, depth, visited, conn, _build_bom_ctx(conn))
        else:
            bom_ctx = _build_bom_ctx(_conn)

    if visited is None: visited = set()
    if parent_id in visited:
        return {'part_id': parent_id, 'error': 'Circular reference', 'children': []}
    visited = visited | {parent_id}

    part = bom_ctx['parts'].get(parent_id)
    if not part:
        return {'part_id': parent_id, 'error': 'Not found', 'children': []}
        
    children_rows = bom_ctx['bom_children'].get(parent_id, [])

    children = []
    for row in children_rows:
        child_tree = build_bom_tree(row['child_id'], depth + 1, visited, _conn, bom_ctx)
        child_tree['bom_qty']  = row['qty']
        child_tree['depth']    = depth + 1
        child_tree['optional'] = (row['qty'] == 0)
        children.append(child_tree)

    return {
        'part_id':    parent_id,
        'plain_desc': part['plain_desc'],
        'type':       part['type'],
        'unit_cost':  part['unit_cost'],
        'labor_hrs':  part['labor_hrs'],
        'children':   children
    }

# ── Cost Rollup ────────────────────────────────────────────────────────────────

def run_rollup(part_id: str, cycle_visited: set = None, memo: dict = None, _conn = None, bom_ctx: dict = None) -> tuple[float, float]:
    if bom_ctx is None:
        if _conn is None:
            with get_conn() as conn:
                return run_rollup(part_id, cycle_visited, memo, conn, _build_bom_ctx(conn))
        else:
            bom_ctx = _build_bom_ctx(_conn)

    if cycle_visited is None: cycle_visited = set()
    if part_id in cycle_visited: return 0.0, 0.0
    
    if memo is not None and part_id in memo:
        return memo[part_id]
        
    cycle_visited = cycle_visited | {part_id}

    part = bom_ctx['parts'].get(part_id)
    if not part: return 0.0, 0.0
    
    if part['type'] in PART_TYPES_STATIC_COST:
        res = float(part['unit_cost'] or 0), float(part['labor_hrs'] or 0)
        if memo is not None: memo[part_id] = res
        return res
        
    children = bom_ctx['bom_children'].get(part_id, [])

    total_mat = 0.0
    total_lbr = float(part['labor_hrs'] or 0)
    
    for child in children:
        if child['qty'] == 0: continue
        c_mat, c_lbr = run_rollup(child['child_id'], cycle_visited, memo, _conn, bom_ctx)
        total_mat += c_mat * child['qty']
        total_lbr += c_lbr * child['qty']
        
    if _conn is not None:
        old_row = _conn.execute("SELECT unit_cost FROM parts WHERE part_id=?", (part_id,)).fetchone()
        old_uc = float(old_row[0] or 0) if old_row else 0.0
        new_uc = round(total_mat, 6)
        _conn.execute("UPDATE parts SET unit_cost=?, rolled_labor_hrs=? WHERE part_id=?",
                     (new_uc, round(total_lbr, 4), part_id))
        if new_uc != old_uc and old_uc > 0:
            _conn.execute("INSERT INTO part_cost_history (part_id, old_cost, new_cost) VALUES (?, ?, ?)", (part_id, old_uc, new_uc))
                     
    res = (total_mat, total_lbr)
    if memo is not None: memo[part_id] = res
    return res
def rollup_all():
    """Re-calculate material unit_cost and rolled_labor_hrs for all FAB/ASSY parts in O(N)."""
    with get_conn() as conn:
        ids = [r[0] for r in conn.execute(
            "SELECT part_id FROM parts WHERE type IN ('FAB','ASSY')"
        ).fetchall()]
        memo = {}
        for pid in ids:
            run_rollup(pid, memo=memo, _conn=conn)

# ── Projects CRUD ──────────────────────────────────────────────────────────────

def get_all_projects() -> list[dict]:
    with get_conn() as conn:
        return [dict(r) for r in conn.execute(
            "SELECT * FROM projects ORDER BY created_at DESC"
        ).fetchall()]

def get_project(project_id: str) -> Optional[dict]:
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM projects WHERE project_id=?", (project_id,)
        ).fetchone()
        return dict(row) if row else None

def upsert_project(data: dict) -> tuple[bool, str]:
    pid = _clean_project_id(data.get('project_id', ''))
    if not pid: return False, "Project ID is required."
    LOGGED_FIELDS = ['status','customer','notes','labor_rate','markup']
    with get_conn() as conn:
        existing = conn.execute(
            "SELECT * FROM projects WHERE project_id=?", (pid,)
        ).fetchone()
        old_row = dict(existing) if existing else None
        vals = (data.get('status','ACTIVE'), data.get('customer',''),
                data.get('notes',''), float(data.get('labor_rate') or 25),
                float(data.get('markup') or 0))
        if old_row:
            conn.execute(
                "UPDATE projects SET status=?,customer=?,notes=?,labor_rate=?,markup=? WHERE project_id=?",
                (*vals, pid))
            new_row = dict(zip(LOGGED_FIELDS, vals))
            _diff_log(conn, 'project', pid, old_row, new_row, LOGGED_FIELDS)
        else:
            conn.execute(
                "INSERT INTO projects (status,customer,notes,labor_rate,markup,project_id) VALUES (?,?,?,?,?,?)",
                (*vals, pid))
            log_change(conn, 'project', pid, 'create')
    return True, pid

def delete_project(project_id: str) -> tuple[bool, str]:
    with get_conn() as conn:
        conn.execute("DELETE FROM projects WHERE project_id=?", (project_id,))
        log_change(conn, 'project', project_id, 'delete')
    return True, "Project deleted."

def add_project_item(project_id: str, part_id: str, qty: float) -> tuple[bool, str]:
    with get_conn() as conn:
        conn.execute(
            "INSERT INTO project_items (project_id,part_id,qty) VALUES (?,?,?)",
            (project_id, part_id, qty))
        log_change(conn, 'project_item', project_id, 'create',
                   field='part_id', old_val=None, new_val=f"{part_id} qty={qty}")
    return True, "Item added."

def update_project_item(item_id: int, qty: float = None, picked: int = None,
                        box_num: str = None, discount_pct: float = None,
                        discount_flat: float = None) -> tuple[bool, str]:
    with get_conn() as conn:
        old = conn.execute("SELECT * FROM project_items WHERE id=?", (item_id,)).fetchone()
        old_row = dict(old) if old else {}
        project_id = old_row.get('project_id', str(item_id))
        part_id    = old_row.get('part_id', str(item_id))
        entity_id  = f"{project_id}/{part_id}"

        if qty          is not None:
            conn.execute("UPDATE project_items SET qty=?          WHERE id=?", (qty,          item_id))
            log_change(conn, 'project_item', entity_id, 'update', 'qty',          old_row.get('qty'),          qty)
        if picked       is not None:
            conn.execute("UPDATE project_items SET picked=?       WHERE id=?", (picked,       item_id))
        if box_num      is not None:
            conn.execute("UPDATE project_items SET box_num=?      WHERE id=?", (box_num,      item_id))
        if discount_pct is not None:
            conn.execute("UPDATE project_items SET discount_pct=? WHERE id=?", (discount_pct, item_id))
            log_change(conn, 'project_item', entity_id, 'update', 'discount_pct', old_row.get('discount_pct'), discount_pct)
        if discount_flat is not None:
            conn.execute("UPDATE project_items SET discount_flat=? WHERE id=?", (discount_flat, item_id))
            log_change(conn, 'project_item', entity_id, 'update', 'discount_flat', old_row.get('discount_flat'), discount_flat)
    return True, "Updated."

def delete_project_item(item_id: int) -> tuple[bool, str]:
    with get_conn() as conn:
        old = conn.execute("SELECT project_id, part_id, qty FROM project_items WHERE id=?", (item_id,)).fetchone()
        conn.execute("DELETE FROM project_items WHERE id=?", (item_id,))
        if old:
            log_change(conn, 'project_item', f"{old['project_id']}/{old['part_id']}", 'delete',
                       field='part_id', old_val=f"{old['part_id']} qty={old['qty']}", new_val=None)
    return True, "Item removed."

def _upsert_part_row(conn, row_dict: dict):
    """Unified upsert logic for parts table imports."""
    part_id = str(row_dict.get('part_id') or '').strip()
    if not part_id: return
    pkg_size  = float(row_dict.get('pkg_size')  or 1)
    pkg_cost  = float(row_dict.get('pkg_cost')  or 0)
    unit_cost = float(row_dict.get('unit_cost') or 0)
    ptype     = str(row_dict.get('type') or '').strip()
    if ptype in PART_TYPES_STATIC_COST and pkg_size > 0 and pkg_cost > 0:
        unit_cost = round(pkg_cost / pkg_size, 6)
    pkg_size_2  = float(row_dict.get('pkg_size_2')  or 1)
    pkg_cost_2  = float(row_dict.get('pkg_cost_2')  or 0)
    unit_cost_2 = float(row_dict.get('unit_cost_2') or 0)
    if ptype in PART_TYPES_STATIC_COST and pkg_size_2 > 0 and pkg_cost_2 > 0:
        unit_cost_2 = round(pkg_cost_2 / pkg_size_2, 6)
    use_alt = int(bool(row_dict.get('use_alt_supplier') and str(row_dict.get('use_alt_supplier')) not in ('0','')))
    
    old_row = conn.execute("SELECT unit_cost FROM parts WHERE part_id=?", (part_id,)).fetchone()
    old_uc = float(old_row[0] or 0) if old_row else 0.0
    
    conn.execute("""INSERT OR REPLACE INTO parts
        (part_id,type,category,base_desc,size_spec,variant,plain_desc,
         supplier,brand_mfg,supplier_pn,uom,pkg_size,pkg_cost,unit_cost,
         labor_hrs,qty_on_hand,cost,on_hand,status,
         supplier_2,brand_mfg_2,supplier_pn_2,pkg_size_2,pkg_cost_2,
         unit_cost_2,use_alt_supplier,last_cost_date)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        part_id, ptype,
        str(row_dict.get('category')    or '').strip(),
        str(row_dict.get('base_desc')   or '').strip(),
        str(row_dict.get('size_spec')   or '').strip(),
        str(row_dict.get('variant')     or '').strip(),
        str(row_dict.get('plain_desc')  or '').strip(),
        str(row_dict.get('supplier')    or '').strip(),
        str(row_dict.get('brand_mfg')   or '').strip(),
        str(row_dict.get('supplier_pn') or '').strip(),
        str(row_dict.get('uom')         or 'ea').strip(),
        pkg_size, pkg_cost, unit_cost,
        float(row_dict.get('labor_hrs')   or 0),
        float(row_dict.get('qty_on_hand') or 0),
        float(row_dict.get('cost')        or 0),
        float(row_dict.get('on_hand')     or 0),
        str(row_dict.get('status') or 'ACTIVE').strip(),
        str(row_dict.get('supplier_2')    or '').strip(),
        str(row_dict.get('brand_mfg_2')   or '').strip(),
        str(row_dict.get('supplier_pn_2') or '').strip(),
        pkg_size_2, pkg_cost_2, unit_cost_2, use_alt,
        str(row_dict.get('last_cost_date') or '').strip(),
    ))
    
    new_uc = unit_cost_2 if (use_alt and unit_cost_2 > 0) else unit_cost
    if old_uc > 0 and new_uc != old_uc:
        conn.execute("INSERT INTO part_cost_history (part_id, old_cost, new_cost) VALUES (?, ?, ?)", (part_id, old_uc, new_uc))

# ── CSV Import ─────────────────────────────────────────────────────────────────

def import_from_csv_data(rows: list[dict]) -> dict:
    results = {'parts': 0, 'bom': 0, 'projects': 0, 'items': 0, 'errors': []}
    with get_conn() as conn:
        for row in rows:
            table = (row.get('_table') or '').strip()
            if table == 'PART':
                try:
                    _upsert_part_row(conn, row)
                    results['parts'] += 1
                except Exception as e:
                    results['errors'].append(f"PART {row.get('part_id')}: {e}")

            elif table == 'BOM':
                try:
                    parent_id = (row.get('parent_id') or '').strip()
                    child_id  = (row.get('child_id')  or '').strip()
                    if not parent_id or not child_id: continue
                    conn.execute(
                        "INSERT OR REPLACE INTO bom (parent_id,child_id,qty) VALUES (?,?,?)",
                        (parent_id, child_id, float(row.get('qty') or 0))
                    )
                    results['bom'] += 1
                except Exception as e:
                    results['errors'].append(f"BOM {row.get('parent_id')}->{row.get('child_id')}: {e}")

            elif table == 'PROJ':
                try:
                    pid = (row.get('project_id') or '').strip()
                    if not pid: continue
                    status_val = (row.get('proj_status') or row.get('status') or 'ACTIVE').strip()
                    conn.execute("""INSERT OR REPLACE INTO projects
                        (project_id,status,customer,notes,labor_rate,markup) VALUES (?,?,?,?,?,?)
                    """, (pid, status_val,
                          (row.get('customer') or '').strip(),
                          (row.get('notes')    or '').strip(),
                          float(row.get('labor_rate') or 25),
                          float(row.get('markup') or 0)))
                    results['projects'] += 1
                except Exception as e:
                    results['errors'].append(f"PROJ {row.get('project_id')}: {e}")

            elif table == 'P_ITEM':
                try:
                    part_id    = (row.get('part_id')    or '').strip()
                    project_id = (row.get('project_id') or '').strip()
                    if not part_id or not project_id: continue
                    itype = (row.get('item_type') or 'ADDITIONAL').strip()
                    if itype not in ('STANDARD','OPTION','ADDITIONAL','DELETED'):
                        itype = 'ADDITIONAL'
                    conn.execute(
                        """INSERT INTO project_items
                           (project_id,part_id,qty,picked,item_type,box_num,discount_pct,discount_flat)
                           VALUES (?,?,?,?,?,?,?,?)""",
                        (project_id, part_id, float(row.get('qty') or 1),
                         int(row.get('picked') or 0), itype,
                         (row.get('box_num') or '').strip(),
                         float(row.get('discount_pct') or 0),
                         float(row.get('discount_flat') or 0))
                    )
                    results['items'] += 1
                except Exception as e:
                    results['errors'].append(f"P_ITEM {row.get('part_id')}: {e}")

            elif table == 'P_BOX':
                try:
                    project_id = (row.get('project_id') or '').strip()
                    box_num    = (row.get('box_num') or '').strip()
                    if not project_id or not box_num: continue
                    conn.execute(
                        """INSERT OR REPLACE INTO project_boxes
                           (project_id, box_num, weight, pallet_num) VALUES (?,?,?,?)""",
                        (project_id, box_num,
                         float(row.get('weight') or 0),
                         (row.get('pallet_num') or '').strip())
                    )
                except Exception as e:
                    results['errors'].append(f"P_BOX {row.get('project_id')}/{row.get('box_num')}: {e}")

            elif table == 'P_PALLET':
                try:
                    project_id = (row.get('project_id') or '').strip()
                    pallet_num = (row.get('pallet_num') or '').strip()
                    if not project_id or not pallet_num: continue
                    conn.execute(
                        """INSERT OR REPLACE INTO project_pallets
                           (project_id, pallet_num, weight, dimensions) VALUES (?,?,?,?)""",
                        (project_id, pallet_num,
                         float(row.get('weight') or 0),
                         (row.get('dimensions') or '').strip())
                    )
                except Exception as e:
                    results['errors'].append(f"P_PALLET {row.get('project_id')}/{row.get('pallet_num')}: {e}")

            elif table == 'P_QUOTE':
                try:
                    project_id = (row.get('project_id') or '').strip()
                    if not project_id: continue
                    # Parse other_items JSON safely
                    try:
                        other_items = json.loads(row.get('other_items') or '[]')
                    except Exception:
                        other_items = []
                    save_quote(project_id, {
                        'version':          int(row.get('version') or 1),
                        'status':           (row.get('quote_status') or 'DRAFT').strip(),
                        'currency':         (row.get('currency') or 'USD').strip(),
                        'overhead_rate':    float(row.get('overhead_rate') or 1.0),
                        'markup_pct':       float(row.get('markup_pct') or 0),
                        'freight_inbound':  float(row.get('freight_inbound') or 0),
                        'freight_outbound': float(row.get('freight_outbound') or 0),
                        'cal_gases_cost':   float(row.get('cal_gases_cost') or 0),
                        'cal_gases_freight':float(row.get('cal_gases_freight') or 0),
                        'training_days':    float(row.get('training_days') or 0),
                        'training_cost':    float(row.get('training_cost') or 0),
                        'training_notes':   (row.get('training_notes') or '').strip(),
                        'discount_pct':     float(row.get('discount_pct') or 0),
                        'discount_flat':    float(row.get('discount_flat') or 0),
                        'discount_note':    (row.get('discount_note') or '').strip(),
                        'internal_notes':   (row.get('internal_notes') or '').strip(),
                        'proforma_header':  (row.get('proforma_header') or '').strip(),
                        'proforma_footer':  (row.get('proforma_footer') or '').strip(),
                        'other_items':      other_items,
                    })
                except Exception as e:
                    results['errors'].append(f"P_QUOTE {row.get('project_id')}: {e}")

            elif table == 'P_OTHER':
                try:
                    project_id = (row.get('project_id') or '').strip()
                    if not project_id: continue
                    conn.execute(
                        """INSERT INTO project_other_items
                           (project_id, description, cost, labor_hrs, apply_markup,
                            box_num, discount_pct, discount_flat, show_on_proforma, sort_order)
                           VALUES (?,?,?,?,?,?,?,?,?,?)""",
                        (project_id,
                         (row.get('description') or '').strip(),
                         float(row.get('cost') or 0),
                         float(row.get('labor_hrs') or 0),
                         int(row.get('apply_markup') or 0),
                         (row.get('box_num') or '').strip(),
                         float(row.get('discount_pct') or 0),
                         float(row.get('discount_flat') or 0),
                         int(row.get('show_on_proforma') or 0),
                         int(row.get('sort_order') or 0))
                    )
                except Exception as e:
                    results['errors'].append(f"P_OTHER {row.get('project_id')}: {e}")

    return results

# ── Master Data Replace ────────────────────────────────────────────────────────

def import_master_data(rows: list[dict]) -> dict:
    """
    Full-replace import for master parts/BOM data from a CSV.
    Wipes ALL existing parts and BOM rows before loading, so renamed
    or deleted part IDs can never ghost back into the database.
    Projects and project_items are left untouched.
    Only PART and BOM rows in the CSV are processed.
    """
    results = {'parts': 0, 'bom': 0, 'projects': 0, 'items': 0, 'errors': [],
               'deleted_parts': 0, 'deleted_bom': 0}
               
    # We use a completely raw connection and disable foreign keys here
    # so that we can freely wipe the parts table even if project_items
    # references old parts. This ensures existing projects don't blow up 
    # the constraint check.
    conn = sqlite3.connect(DB_PATH)
    try:
        conn.execute("PRAGMA foreign_keys = OFF")
        conn.execute("BEGIN TRANSACTION")

        # Count what we're about to remove so the UI can report it
        results['deleted_parts'] = conn.execute("SELECT COUNT(*) FROM parts").fetchone()[0]
        results['deleted_bom']   = conn.execute("SELECT COUNT(*) FROM bom").fetchone()[0]

        conn.execute("DELETE FROM bom")
        conn.execute("DELETE FROM parts")

        for row in rows:
            table = (row.get('_table') or '').strip()
            if table == 'PART':
                try:
                    _upsert_part_row(conn, row)
                    results['parts'] += 1
                except Exception as e:
                    results['errors'].append(f"PART {row.get('part_id')}: {e}")

            elif table == 'BOM':
                try:
                    parent_id = (row.get('parent_id') or '').strip()
                    child_id  = (row.get('child_id')  or '').strip()
                    if not parent_id or not child_id: continue
                    conn.execute(
                        "INSERT INTO bom (parent_id,child_id,qty) VALUES (?,?,?)",
                        (parent_id, child_id, float(row.get('qty') or 0))
                    )
                    results['bom'] += 1
                except Exception as e:
                    results['errors'].append(f"BOM {row.get('parent_id')}->{row.get('child_id')}: {e}")

        conn.execute("COMMIT")
    except Exception as e:
        conn.execute("ROLLBACK")
        results['errors'].append(f"Fatal Import Error: {str(e)}")
    finally:
        conn.close()

    return results

# ── SQLite DB direct import ────────────────────────────────────────────────────

def import_from_sqlite(source_path: str) -> dict:
    """Import directly from another LEMS .db file."""
    results = {'parts': 0, 'bom': 0, 'projects': 0, 'items': 0, 'errors': []}
    try:
        src = sqlite3.connect(source_path)
        src.row_factory = sqlite3.Row
    except Exception as e:
        results['errors'].append(f"Cannot open source DB: {e}")
        return results

    with get_conn() as dst:
        # Parts
        try:
            for r in src.execute("SELECT * FROM parts").fetchall():
                r = dict(r)
                _upsert_part_row(dst, r)
                results['parts'] += 1
        except Exception as e:
            results['errors'].append(f"Parts import error: {e}")

        # BOM — skip rows with blank parent (root markers)
        try:
            for r in src.execute("SELECT * FROM bom").fetchall():
                r = dict(r)
                parent_id = (r.get('parent_id') or '').strip()
                child_id  = (r.get('child_id')  or '').strip()
                if not parent_id or not child_id: continue
                dst.execute(
                    "INSERT OR REPLACE INTO bom (parent_id,child_id,qty) VALUES (?,?,?)",
                    (parent_id, child_id, float(r.get('qty') or 0))
                )
                results['bom'] += 1
        except Exception as e:
            results['errors'].append(f"BOM import error: {e}")

        # Projects
        try:
            for r in src.execute("SELECT * FROM projects").fetchall():
                r = dict(r)
                pid = (r.get('project_id') or '').strip()
                if not pid: continue
                dst.execute("""INSERT OR REPLACE INTO projects
                    (project_id,status,customer,notes,labor_rate,markup) VALUES (?,?,?,?,?,?)
                """, (pid, r.get('status','ACTIVE'), r.get('customer','') or '',
                      r.get('notes','') or '', float(r.get('labor_rate') or 25),
                      float(r.get('markup') or 0)))
                results['projects'] += 1
        except Exception as e:
            results['errors'].append(f"Projects import error: {e}")

        # Project items
        try:
            for r in src.execute("SELECT * FROM project_items").fetchall():
                r = dict(r)
                part_id    = (r.get('part_id')    or '').strip()
                project_id = (r.get('project_id') or '').strip()
                if not part_id or not project_id: continue
                itype = (r.get('item_type') or 'ADDITIONAL').strip()
                if itype not in ('STANDARD','OPTION','ADDITIONAL','DELETED'):
                    itype = 'ADDITIONAL'
                dst.execute(
                    "INSERT INTO project_items (project_id,part_id,qty,picked,item_type) VALUES (?,?,?,?,?)",
                    (project_id, part_id, float(r.get('qty') or 1),
                     int(r.get('picked') or 0), itype)
                )
                results['items'] += 1
        except Exception as e:
            results['errors'].append(f"Project items error: {e}")

    # Initialise sort_order after import
    with get_conn() as conn:
        conn.execute("""
            UPDATE bom SET sort_order = rowid
            WHERE sort_order = 0 OR sort_order IS NULL
        """)

    src.close()
    return results

# ── XLSX import ────────────────────────────────────────────────────────────────

def import_from_xlsx(xlsx_path: str) -> dict:
    """Import from an Excel workbook with sheets: parts, bom, projects, project_items."""
    try:
        import openpyxl
    except ImportError:
        return {'parts':0,'bom':0,'projects':0,'items':0,
                'errors':['openpyxl not installed. Run: pip install openpyxl']}

    results = {'parts': 0, 'bom': 0, 'projects': 0, 'items': 0, 'errors': []}
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        results['errors'].append(f"Cannot open xlsx: {e}")
        return results

    def sheet_to_dicts(ws):
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [str(h).strip() if h is not None else '' for h in next(rows)]
        except StopIteration:
            return []
        out = []
        for row in rows:
            if all(v is None for v in row): continue
            out.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
        return out

    with get_conn() as dst:
        if 'parts' in wb.sheetnames:
            for r in sheet_to_dicts(wb['parts']):
                try:
                    _upsert_part_row(dst, r)
                    results['parts'] += 1
                except Exception as e:
                    results['errors'].append(f"PART {r.get('part_id')}: {e}")

        if 'bom' in wb.sheetnames:
            for r in sheet_to_dicts(wb['bom']):
                try:
                    parent_id = str(r.get('parent_id') or '').strip()
                    child_id  = str(r.get('child_id')  or '').strip()
                    if not parent_id or not child_id or parent_id == 'None': continue
                    dst.execute(
                        "INSERT OR REPLACE INTO bom (parent_id,child_id,qty) VALUES (?,?,?)",
                        (parent_id, child_id, float(r.get('qty') or 0))
                    )
                    results['bom'] += 1
                except Exception as e:
                    results['errors'].append(f"BOM {r.get('parent_id')}->{r.get('child_id')}: {e}")

        if 'projects' in wb.sheetnames:
            for r in sheet_to_dicts(wb['projects']):
                try:
                    pid = str(r.get('project_id') or '').strip()
                    if not pid or pid == 'None': continue
                    dst.execute("""INSERT OR REPLACE INTO projects
                        (project_id,status,customer,notes,labor_rate,markup) VALUES (?,?,?,?,?,?)
                    """, (pid, str(r.get('status') or 'ACTIVE'),
                          str(r.get('customer') or ''), str(r.get('notes') or ''),
                          float(r.get('labor_rate') or 25), float(r.get('markup') or 0)))
                    results['projects'] += 1
                except Exception as e:
                    results['errors'].append(f"PROJ {r.get('project_id')}: {e}")

        if 'project_items' in wb.sheetnames:
            for r in sheet_to_dicts(wb['project_items']):
                try:
                    part_id    = str(r.get('part_id')    or '').strip()
                    project_id = str(r.get('project_id') or '').strip()
                    if not part_id or not project_id or part_id == 'None': continue
                    dst.execute(
                        "INSERT INTO project_items (project_id,part_id,qty,picked) VALUES (?,?,?,?)",
                        (project_id, part_id, float(r.get('qty') or 1), int(r.get('picked') or 0))
                    )
                    results['items'] += 1
                except Exception as e:
                    results['errors'].append(f"P_ITEM {r.get('part_id')}: {e}")

    with get_conn() as conn:
        conn.execute("UPDATE bom SET sort_order=rowid WHERE sort_order=0 OR sort_order IS NULL")

    wb.close()
    return results

# ── Where-used (full ancestor paths) ──────────────────────────────────────────

def get_where_used(part_id: str) -> list[dict]:
    results = []
    with get_conn() as conn:
        bom_ctx = _build_bom_ctx(conn)
        _collect_where_used(conn, part_id, [], results, set(), bom_ctx)
    return results

def _collect_where_used(conn, part_id: str, current_path: list, results: list, visited: set, bom_ctx: dict = None):
    if bom_ctx is None:
        bom_ctx = _build_bom_ctx(conn)
    if part_id in visited:
        return
    visited = visited | {part_id}
    parents = bom_ctx['bom_parents'].get(part_id, [])
    for par_row in parents:
        parent_id = par_row['parent_id']
        part = bom_ctx['parts'].get(parent_id, {})
        path = [parent_id] + current_path
        results.append({
            'parent_id':   parent_id,
            'parent_desc': part.get('plain_desc', ''),
            'parent_type': part.get('type', ''),
            'qty':         par_row['qty'],
            'path':        path,
        })
        _collect_where_used(conn, parent_id, path, results, visited, bom_ctx)

def get_direct_parents(part_id: str) -> list[dict]:
    """Direct parents only, with qty."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT b.parent_id, b.qty, p.plain_desc, p.type
            FROM bom b JOIN parts p ON p.part_id = b.parent_id
            WHERE b.child_id = ?
            ORDER BY p.type, b.parent_id
        """, (part_id,)).fetchall()
    return [dict(r) for r in rows]

# ── BOM flat explode ───────────────────────────────────────────────────────────

def explode_bom_flat(parent_id: str, qty_mult: float = 1.0,
                     visited: set = None, _conn = None, bom_ctx: dict = None) -> list[dict]:
    if bom_ctx is None:
        if _conn is None:
            with get_conn() as conn:
                return explode_bom_flat(parent_id, qty_mult, visited, conn, _build_bom_ctx(conn))
        else:
            bom_ctx = _build_bom_ctx(_conn)

    if visited is None: visited = set()
    if parent_id in visited: return []
    visited = visited | {parent_id}

    children = bom_ctx['bom_children'].get(parent_id, [])

    rows = []
    for child in children:
        child_id = child['child_id']
        p = bom_ctx['parts'].get(child_id)
        if not p: continue
        qty = child['qty']
        optional = (qty == 0)
        eff_qty  = qty * qty_mult if not optional else 0

        rows.append({
            'part_id':    child_id,
            'plain_desc': p.get('plain_desc', ''),
            'type':       p.get('type', ''),
            'unit_cost':  p.get('unit_cost', 0),
            'uom':        p.get('uom', 'ea'),
            'total_qty':  eff_qty,
            'optional':   optional,
            'status':     p.get('status', 'ACTIVE')
        })
        if p.get('type') in PART_TYPES_WITH_BOM:
            sub = explode_bom_flat(child_id, eff_qty, visited, _conn, bom_ctx)
            rows.extend(sub)

    return rows
def explode_bom_flat_deduped(parent_id: str) -> list[dict]:
    """
    Like explode_bom_flat but collapses duplicate part IDs into a single row,
    summing total_qty. Also fetches supplier info for each part.
    Optional parts (qty=0) are collected separately and appended at the end.
    Used for the printed BOM flat parts list.
    """
    raw = explode_bom_flat(parent_id)

    # Merge by part_id
    seen: dict = {}
    for row in raw:
        pid = row['part_id']
        if pid not in seen:
            seen[pid] = dict(row)
        else:
            if not row['optional']:
                seen[pid]['total_qty'] = (seen[pid]['total_qty'] or 0) + (row['total_qty'] or 0)
            # If we already have it as required, keep it required
            if not row['optional']:
                seen[pid]['optional'] = False

    # Enrich with supplier info
    with get_conn() as conn:
        for pid, row in seen.items():
            part = conn.execute(
                "SELECT supplier, supplier_pn, brand_mfg, use_alt_supplier, "
                "supplier_2, supplier_pn_2, brand_mfg_2 FROM parts WHERE part_id=?",
                (pid,)
            ).fetchone()
            if part:
                use_alt = part['use_alt_supplier']
                row['supplier']    = part['supplier_2']    if use_alt and part['supplier_2']    else part['supplier']
                row['supplier_pn'] = part['supplier_pn_2'] if use_alt and part['supplier_pn_2'] else part['supplier_pn']
                row['brand_mfg']   = part['brand_mfg_2']   if use_alt and part['brand_mfg_2']   else part['brand_mfg']
            else:
                row['supplier'] = row['supplier_pn'] = row['brand_mfg'] = ''

    # Sort: required first (by part_id), then optional
    required = sorted([r for r in seen.values() if not r['optional']], key=lambda x: x['part_id'])
    optional = sorted([r for r in seen.values() if  r['optional']],    key=lambda x: x['part_id'])
    return required + optional


# ── Print / PDF data builders ─────────────────────────────────────────────────

def build_print_project(project_id: str) -> Optional[dict]:
    """
    Assemble print data using get_project_summary for full consistency
    with the order sheet — same math, same DELETED handling, same categories.
    """
    ps = get_project_summary(project_id)
    if not ps: return None
    # Attach BOM tree to each line item for the breakdown section
    for item in ps['line_items']:
        item['tree'] = build_bom_tree(item['part_id'])
    ps['grand_total'] = ps['retail']
    ps['markup']      = float(ps['project'].get('markup') or 0)
    return ps


def _migrate_other_items(conn):
    """One-time migration: move project_quotes.other_items JSON into project_other_items table."""
    try:
        quotes = conn.execute(
            "SELECT project_id, other_items FROM project_quotes WHERE other_items IS NOT NULL AND other_items != '[]'"
        ).fetchall()
        for q in quotes:
            existing = conn.execute(
                "SELECT COUNT(*) FROM project_other_items WHERE project_id=?", (q['project_id'],)
            ).fetchone()[0]
            if existing > 0:
                continue
            items = json.loads(q['other_items'] or '[]')
            for idx, i in enumerate(items):
                if not isinstance(i, dict): continue
                conn.execute(
                    """INSERT INTO project_other_items
                       (project_id, description, cost, labor_hrs, apply_markup,
                        box_num, discount_pct, discount_flat, show_on_proforma, sort_order)
                       VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    (q['project_id'],
                     (i.get('desc') or i.get('description') or '').strip(),
                     float(i.get('cost') or 0),
                     float(i.get('labor_hrs') or 0),
                     1 if i.get('apply_markup') else 0,
                     (i.get('box_num') or '').strip(),
                     float(i.get('discount_pct') or 0),
                     float(i.get('discount_flat') or 0),
                     1 if i.get('show_on_proforma') else 0,
                     idx)
                )
    except Exception:
        pass  # Non-critical migration, never block startup

# ── Material cost + labor hours rollup (no labor rate applied) ────────────────

def calc_bom_summary(part_id: str, qty_mult: float = 1.0,
                     visited: set = None) -> dict:
    """
    Returns {material_cost, labor_hrs} for an assembly in O(1) time by 
    reading pre-rolled values from the parts table.
    """
    with get_conn() as conn:
        part = conn.execute(
            "SELECT type, unit_cost, labor_hrs, rolled_labor_hrs FROM parts WHERE part_id=?", 
            (part_id,)
        ).fetchone()
        
    if not part:
        return {'material_cost': 0.0, 'labor_hrs': 0.0}

    # unit_cost is ALWAYS pure material cost.
    mat_cost = float(part['unit_cost'] or 0) * qty_mult

    if part['type'] in PART_TYPES_STATIC_COST:
        lbr_hrs = float(part['labor_hrs'] or 0) * qty_mult
    else:
        lbr_hrs = float(part['rolled_labor_hrs'] or 0) * qty_mult

    return {'material_cost': round(mat_cost, 6), 'labor_hrs': round(lbr_hrs, 4)}

# ── Flags and warnings ────────────────────────────────────────────────────────

def get_system_flags() -> dict:
    """Return categorised warnings for the admin/dashboard flags panel."""
    flags = {
        'zero_cost':     [],   # PRT/RAW with unit_cost = 0
        'empty_bom':     [],   # FAB/ASSY with no BOM children
        'orphaned':      [],   # PRT/RAW not used in any BOM
        'obsolete_bom':  [],   # OBSOLETE parts still in active BOMs
        'missing_desc':  [],   # parts with no plain_desc
        'missing_supplier': [], # PRT/RAW with no supplier
    }
    with get_conn() as conn:
        # Zero-cost purchased parts
        for r in conn.execute("""
            SELECT part_id, type, category, plain_desc FROM parts
            WHERE type IN ('PRT','RAW') AND (unit_cost IS NULL OR unit_cost = 0)
            AND (status IS NULL OR status != 'OBSOLETE')
            ORDER BY category, part_id
        """).fetchall():
            flags['zero_cost'].append(dict(r))

        # FAB/ASSY with no BOM children
        for r in conn.execute("""
            SELECT p.part_id, p.type, p.category, p.plain_desc FROM parts p
            WHERE p.type IN ('FAB','ASSY')
            AND p.part_id NOT IN (
                SELECT DISTINCT parent_id FROM bom
                WHERE parent_id IS NOT NULL AND parent_id != ''
            )
            AND (p.status IS NULL OR p.status != 'OBSOLETE')
            ORDER BY p.type, p.part_id
        """).fetchall():
            flags['empty_bom'].append(dict(r))

        # PRT/RAW not used as a child in any BOM
        for r in conn.execute("""
            SELECT p.part_id, p.type, p.category, p.plain_desc FROM parts p
            WHERE p.type IN ('PRT','RAW')
            AND p.part_id NOT IN (SELECT DISTINCT child_id FROM bom)
            AND (p.status IS NULL OR p.status != 'OBSOLETE')
            ORDER BY p.category, p.part_id
        """).fetchall():
            flags['orphaned'].append(dict(r))

        # Obsolete parts still referenced in BOMs — one entry per unique part
        for r in conn.execute("""
            SELECT b.child_id as part_id, p.type, p.category, p.plain_desc,
                   COUNT(DISTINCT b.parent_id) as parent_count
            FROM bom b JOIN parts p ON p.part_id = b.child_id
            WHERE p.status = 'OBSOLETE'
            GROUP BY b.child_id
            ORDER BY b.child_id
        """).fetchall():
            row = dict(r)
            row['parent_id'] = None  # not meaningful for multi-parent case
            flags['obsolete_bom'].append(row)

        # Missing description
        for r in conn.execute("""
            SELECT part_id, type, category FROM parts
            WHERE plain_desc IS NULL OR plain_desc = ''
            ORDER BY type, part_id
        """).fetchall():
            flags['missing_desc'].append(dict(r))

        # PRT/RAW missing supplier
        for r in conn.execute("""
            SELECT part_id, type, category, plain_desc FROM parts
            WHERE type IN ('PRT','RAW')
            AND (supplier IS NULL OR supplier = '')
            AND (status IS NULL OR status != 'OBSOLETE')
            AND unit_cost > 0
            ORDER BY category, part_id
        """).fetchall():
            flags['missing_supplier'].append(dict(r))

    return flags

# ── Pick list generator ────────────────────────────────────────────────────────

def generate_pick_list(project_id: str) -> list[dict]:
    """
    Explode all project line items, subtract DELETED contributions.
    Returns sorted unique PRT/RAW with net total_qty >= 0.
    Includes 'picked' status from project_pick_status table.
    """
    with get_conn() as conn:
        items = conn.execute(
            "SELECT pi.part_id, pi.qty, pi.item_type, p.plain_desc, p.type, p.unit_cost, p.uom, p.status "
            "FROM project_items pi JOIN parts p ON p.part_id = pi.part_id "
            "WHERE pi.project_id=?",
            (project_id,)
        ).fetchall()
        
        pick_status = {
            r['part_id']: {
                'picked': bool(r['picked']),
                'picked_qty': float(r['picked_qty'] or 0.0)
            }
            for r in conn.execute(
                "SELECT part_id, picked, picked_qty FROM project_pick_status WHERE project_id=?",
                (project_id,)
            ).fetchall()
        }

    aggregated: dict[str, float] = {}
    meta: dict[str, dict] = {}

    def _add_part(p: dict, qty: float, sign: float):
        if p['type'] not in ('PRT','RAW'): return
        pid = p['part_id']
        aggregated[pid] = aggregated.get(pid, 0.0) + sign * qty
        if pid not in meta:
            meta[pid] = {'part_id': pid, 'plain_desc': p['plain_desc'],
                         'type': p['type'], 'unit_cost': p['unit_cost'], 'uom': p['uom']}

    with get_conn() as conn:
        for item in items:
            sign = -1.0 if item['item_type'] == 'DELETED' else 1.0
            p = dict(item)
            
            if p['type'] in ('PRT','RAW'):
                _add_part(p, float(item['qty']), sign)
            else:
                flat = explode_bom_flat(item['part_id'], qty_mult=float(item['qty']), _conn=conn)
                for row in flat:
                    if row['type'] not in ('PRT','RAW'): continue
                    if row['optional'] or row['total_qty'] == 0: continue
                    _add_part(row, row['total_qty'], sign)

    # Enrich with inventory data, clamp negatives to 0
    result = []
    with get_conn() as conn:
        for pid in sorted(aggregated, key=lambda x: (meta.get(x,{}).get('type',''), x)):
            net_qty = max(0.0, aggregated[pid])
            if net_qty == 0 and aggregated[pid] >= 0:
                continue
            row = meta.get(pid, {})
            if not row:
                continue
            p = conn.execute("""
                SELECT qty_on_hand, qty_on_order, order_eta,
                       supplier, supplier_pn, category
                FROM parts WHERE part_id=?
            """, (pid,)).fetchone()
            on_hand  = float(p['qty_on_hand']  or 0) if p else 0
            on_order = float(p['qty_on_order'] or 0) if p else 0
            shortage = max(0.0, net_qty - on_hand - on_order)
            ps = pick_status.get(pid, {})
            result.append({
                **row,
                'total_qty':  net_qty,
                'on_hand':    on_hand,
                'on_order':   on_order,
                'order_eta':  (p['order_eta'] or '') if p else '',
                'shortage':   shortage,
                'ext_cost':   row.get('unit_cost', 0) * net_qty,
                'supplier':   p['supplier']    if p else '',
                'supplier_pn':p['supplier_pn'] if p else '',
                'category':   p['category']    if p else '',
                'picked':     ps.get('picked', False),
                'picked_qty': ps.get('picked_qty', 0.0),
            })

    return result

# ── Project item type management ───────────────────────────────────────────────

ITEM_TYPES = ['STANDARD', 'OPTION', 'ADDITIONAL', 'DELETED']

def set_project_item_type(item_id: int, item_type: str) -> tuple[bool, str]:
    if item_type not in ITEM_TYPES:
        return False, f"Invalid type: {item_type}"
    with get_conn() as conn:
        old = conn.execute("SELECT project_id, part_id, item_type FROM project_items WHERE id=?", (item_id,)).fetchone()
        conn.execute("UPDATE project_items SET item_type=? WHERE id=?", (item_type, item_id))
        if old:
            log_change(conn, 'project_item', f"{old['project_id']}/{old['part_id']}", 'update',
                       field='item_type', old_val=old['item_type'], new_val=item_type)
    return True, "Updated."

def get_project_items(project_id: str) -> list[dict]:
    with get_conn() as conn:
        return [dict(r) for r in conn.execute("""
            SELECT pi.id, pi.part_id, pi.qty, pi.picked,
                   pi.item_type, pi.box_num, pi.discount_pct, pi.discount_flat,
                   p.plain_desc, p.type, p.unit_cost, p.uom, p.status
            FROM project_items pi JOIN parts p ON p.part_id=pi.part_id
            WHERE pi.project_id=?
            ORDER BY
                CASE pi.item_type
                    WHEN 'STANDARD'   THEN 1
                    WHEN 'OPTION'     THEN 2
                    WHEN 'ADDITIONAL' THEN 3
                    WHEN 'DELETED'    THEN 4
                    ELSE 5 END,
                p.type, pi.part_id
        """, (project_id,)).fetchall()]

# ── Inventory helpers ─────────────────────────────────────────────────────────

def update_inventory(part_id: str, qty_on_hand: float = None,
                     qty_on_order: float = None,
                     order_eta: str = None) -> tuple[bool, str]:
    with get_conn() as conn:
        logs = []
        if qty_on_hand  is not None:
            conn.execute("UPDATE parts SET qty_on_hand=?  WHERE part_id=?", (qty_on_hand,  part_id))
            logs.append(f"OH:{qty_on_hand}")
        if qty_on_order is not None:
            conn.execute("UPDATE parts SET qty_on_order=? WHERE part_id=?", (qty_on_order, part_id))
            logs.append(f"OO:{qty_on_order}")
        if order_eta    is not None:
            conn.execute("UPDATE parts SET order_eta=?    WHERE part_id=?", (order_eta,    part_id))
            logs.append(f"ETA:{order_eta}")
        if logs:
            log_change(conn, 'part', part_id, 'UPDATE_INVENTORY', new_val=" ".join(logs))
    return True, "Inventory updated."

def commit_picks(project_id: str, picks: list[dict]) -> tuple[bool, str]:
    """
    Persist pick selections in project_pick_status and adjust quantities securely.
    ONLY affects the parts included in `picks`. Unmentioned parts are left as-is.
    """
    try:
        with get_conn() as conn:
            for p in picks:
                part_id = p['part_id']
                new_qty = float(p.get('qty', 0))

                # Check current status for this specific part
                row = conn.execute(
                    "SELECT picked_qty FROM project_pick_status WHERE project_id=? AND part_id=?", 
                    (project_id, part_id)
                ).fetchone()

                old_qty = 0.0
                if row:
                    old_qty = float(row['picked_qty'] or 0.0)
                    if old_qty == 0:
                        # Legacy fallback: assume current new_qty was the original picked amount
                        if new_qty > 0:
                            old_qty = new_qty
                        else:
                            # If unpicking a legacy item, we don't know the amount. Default to simply not refunding it.
                            pass

                delta = new_qty - old_qty

                # Update on_hand inventory if there's a difference
                if delta != 0:
                    p_row = conn.execute(
                        "SELECT qty_on_hand FROM parts WHERE part_id=?", (part_id,)
                    ).fetchone()
                    if p_row:
                        current_oh = float(p_row['qty_on_hand'] or 0)
                        conn.execute(
                            "UPDATE parts SET qty_on_hand=? WHERE part_id=?",
                            (current_oh - delta, part_id)
                        )

                # Update the project_pick_status table
                if new_qty > 0:
                    conn.execute(
                        """INSERT INTO project_pick_status (project_id, part_id, picked, picked_qty)
                           VALUES (?, ?, 1, ?)
                           ON CONFLICT(project_id, part_id) DO UPDATE SET
                           picked=1, picked_qty=excluded.picked_qty""",
                        (project_id, part_id, new_qty)
                    )
                else:
                    conn.execute(
                        "DELETE FROM project_pick_status WHERE project_id=? AND part_id=?",
                        (project_id, part_id)
                    )
                log_change(conn, 'project', project_id, 'COMMIT_PICK', field=part_id, old_val=old_qty, new_val=new_qty)

        return True, "Picks applied and inventory updated."
    except Exception as e:
        return False, str(e)

def get_inventory_list(search: str = '', cat_filter: str = '',
                       show_shortages_only: bool = False) -> list[dict]:
    with get_conn() as conn:
        q = """
            SELECT part_id, type, category, plain_desc, supplier, supplier_pn,
                   uom, unit_cost, qty_on_hand,
                   COALESCE(qty_on_order, 0) as qty_on_order,
                   COALESCE(order_eta, '')   as order_eta
            FROM parts
            WHERE type IN ('PRT','RAW')
            AND (status IS NULL OR status != 'OBSOLETE')
        """
        params = []
        if search:
            q += " AND (part_id LIKE ? OR plain_desc LIKE ? OR supplier LIKE ?)"
            s = f'%{search}%'
            params += [s, s, s]
        if cat_filter:
            q += " AND category = ?"
            params.append(cat_filter)
        q += " ORDER BY category, part_id"
        rows = [dict(r) for r in conn.execute(q, params).fetchall()]

    if show_shortages_only:
        rows = [r for r in rows if r['qty_on_hand'] <= 0 and r['unit_cost'] > 0]
    return rows

# ── v2.4 additions ─────────────────────────────────────────────────────────────

def _clean_project_id(text: str) -> str:
    """Project IDs allow alphanumeric, hyphens, and underscores."""
    cleaned = re.sub(r'[^A-Z0-9_-]', '', str(text).upper().strip()); return cleaned

def clone_project(source_id: str, new_id: str) -> tuple[bool, str]:
    """Duplicate a project with all its line items into a new project_id."""
    new_id = _clean_project_id(new_id)
    if not new_id:
        return False, "New project ID is required."
    source = get_project(source_id)
    if not source:
        return False, f"Source project '{source_id}' not found."
    with get_conn() as conn:
        existing = conn.execute(
            "SELECT project_id FROM projects WHERE project_id=?", (new_id,)
        ).fetchone()
        if existing:
            return False, f"Project '{new_id}' already exists."
        # Clone project row — clear customer/notes, keep rates
        conn.execute("""
            INSERT INTO projects (project_id, status, customer, notes, labor_rate, markup)
            VALUES (?,?,?,?,?,?)
        """, (new_id, 'ACTIVE', '', '', source['labor_rate'], source['markup']))
        log_change(conn, 'project', new_id, 'CLONE_PROJECT', old_val=source_id)

        # Clone line items including box_num and discounts
        items = conn.execute(
            """SELECT part_id, qty, item_type,
               COALESCE(box_num,'') as box_num,
               COALESCE(discount_pct,0) as discount_pct,
               COALESCE(discount_flat,0) as discount_flat
               FROM project_items WHERE project_id=?""",
            (source_id,)
        ).fetchall()
        for item in items:
            conn.execute(
                """INSERT INTO project_items
                   (project_id, part_id, qty, item_type, box_num, discount_pct, discount_flat)
                   VALUES (?,?,?,?,?,?,?)""",
                (new_id, item['part_id'], item['qty'],
                 item['item_type'] or 'ADDITIONAL',
                 item['box_num'], item['discount_pct'], item['discount_flat'])
            )

        # Clone packing metadata
        boxes = conn.execute(
            "SELECT box_num, weight, pallet_num FROM project_boxes WHERE project_id=?",
            (source_id,)
        ).fetchall()
        for b in boxes:
            conn.execute(
                "INSERT INTO project_boxes (project_id, box_num, weight, pallet_num) VALUES (?,?,?,?)",
                (new_id, b['box_num'], b['weight'], b['pallet_num'])
            )
        pallets = conn.execute(
            "SELECT pallet_num, weight, dimensions FROM project_pallets WHERE project_id=?",
            (source_id,)
        ).fetchall()
        for p in pallets:
            conn.execute(
                "INSERT INTO project_pallets (project_id, pallet_num, weight, dimensions) VALUES (?,?,?,?)",
                (new_id, p['pallet_num'], p['weight'], p['dimensions'])
            )

    # Clone quote — carry over ALL settings including additional costs and quoted labor rate
    src_quote = get_or_create_quote(source_id)
    cloned_other = src_quote.get('other_items') or []
    save_quote(new_id, {
        'overhead_rate':    src_quote.get('overhead_rate', 1.0),
        'markup_pct':       src_quote.get('markup_pct', 0),
        'labor_rate_quoted': src_quote.get('labor_rate_quoted', 0),
        'currency':         src_quote.get('currency', 'USD'),
        'other_items':      cloned_other,
        'proforma_header':  src_quote.get('proforma_header', ''),
        'proforma_footer':  src_quote.get('proforma_footer', ''),
        'internal_notes':   '',
        'status':           src_quote.get('status', 'DRAFT'),
        'version':          src_quote.get('version', 1),
        # Carry over additional cost line items
        'freight_inbound':  src_quote.get('freight_inbound', 0),
        'freight_outbound': src_quote.get('freight_outbound', 0),
        'cal_gases_cost':   src_quote.get('cal_gases_cost', 0),
        'cal_gases_freight':src_quote.get('cal_gases_freight', 0),
        'training_days':    src_quote.get('training_days', 0),
        'training_cost':    src_quote.get('training_cost', 0),
        'training_notes':   src_quote.get('training_notes', ''),
        # Reset discount and freeze — those are deal-specific
        'discount_pct': 0, 'discount_flat': 0, 'discount_note': '',
        'costs_frozen': 0, 'frozen_at': '',
        'frozen_material': 0, 'frozen_labor_hrs': 0,
    })

    # Clone custom/quoted items from project_other_items table
    src_items = get_project_other_items(source_id)
    if src_items:
        with get_conn() as conn:
            for oi in src_items:
                conn.execute(
                    """INSERT INTO project_other_items
                       (project_id, description, cost, labor_hrs, apply_markup,
                        box_num, discount_pct, discount_flat, show_on_proforma, sort_order)
                       VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    (new_id, oi['description'], oi['cost'], oi['labor_hrs'],
                     oi['apply_markup'], oi['box_num'], oi['discount_pct'],
                     oi['discount_flat'], oi['show_on_proforma'], oi['sort_order'])
                )

    return True, new_id

def get_parent_counts() -> dict[str, int]:
    """Return {part_id: parent_count} for every part that has at least one parent."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT child_id, COUNT(DISTINCT parent_id) as cnt
            FROM bom
            WHERE parent_id IS NOT NULL AND parent_id != ''
            GROUP BY child_id
        """).fetchall()
    return {r['child_id']: r['cnt'] for r in rows}

def run_rollup_for_part_and_ancestors(part_id: str):
    """
    After a BOM change under `part_id`, roll up that assembly plus
    every ancestor that depends on it — so costs stay current immediately.

    Ordering note: BFS starting from `part_id` and walking upward produces
    `to_rollup` in leaf-to-root order (changed node first, then its parents,
    then grandparents, etc.). This is the correct direction for rollup:
    `run_rollup` for a parent reads its children's already-stored `unit_cost`,
    so children must be rolled up before their parents. Because we seed BFS
    from the changed node — which is already the deepest point of change —
    the list is naturally in the right order without an explicit topological sort.

    `run_rollup` is also memoized via the shared `memo` dict, so any
    sub-assembly that appears in multiple ancestor paths is only computed once.
    """
    with get_conn() as conn:
        # Collect all ancestors (breadth-first, leaf → root order)
        to_rollup = [part_id]
        visited   = {part_id}
        queue     = [part_id]
        while queue:
            node = queue.pop(0)
            parents = conn.execute(
                "SELECT DISTINCT parent_id FROM bom WHERE child_id=? AND parent_id!=''",
                (node,)
            ).fetchall()
            for p in parents:
                pid = p['parent_id']
                if pid not in visited:
                    visited.add(pid)
                    to_rollup.append(pid)
                    queue.append(pid)

        memo = {}
        for pid in to_rollup:
            p = _get_part(conn, pid)
            if p and p['type'] in PART_TYPES_WITH_BOM:
                run_rollup(pid, memo=memo, _conn=conn)

# ── v2.5 additions ─────────────────────────────────────────────────────────────

# Standardized UOM values
UOM_OPTIONS = ['ea', 'ft', 'in', 'hr', 'lb', 'oz', 'kg', 'ml', 'L', 'kit', 'lot', 'set', 'pr', 'box', 'roll']

def rename_part(old_id: str, new_id: str) -> tuple[bool, str]:
    """
    Permanently rename a part_id globally.
    SQLite ON UPDATE CASCADE automatically handles updating BOM, project_items, and project_pick_status.
    """
    new_id = build_part_id(
        new_id.split('-')[0] if '-' in new_id else new_id,
        new_id.split('-')[1] if new_id.count('-') >= 1 else '',
        new_id.split('-')[2] if new_id.count('-') >= 2 else '',
        '-'.join(new_id.split('-')[3:-1]) if new_id.count('-') >= 3 else '',
        new_id.split('-')[-1] if new_id.count('-') >= 4 else '',
    ) if '-' not in new_id else clean(new_id)  # accept fully-formed IDs directly

    if not new_id or new_id == old_id:
        return False, "New part ID is the same as current or invalid."
    with get_conn() as conn:
        existing = conn.execute("SELECT part_id FROM parts WHERE part_id=?", (new_id,)).fetchone()
        if existing:
            return False, f"Part ID '{new_id}' already exists."
        
        # We rely entirely on the ON UPDATE CASCADE constraints on the bom, project_items, and project_pick_status tables.
        # This single update will automatically and safely propagate the new ID.
        conn.execute("UPDATE parts SET part_id=? WHERE part_id=?", (new_id, old_id))
        
        log_change(conn, 'part', new_id, 'RENAME', '', old_id, new_id)
        
    return True, new_id

def get_count_sheet_data(project_id: str = None, cat_filter: str = '') -> list[dict]:
    """
    Data for a physical inventory count sheet.
    If project_id given: only parts in that project's pick list.
    Otherwise: all PRT/RAW parts (optionally filtered by category).
    Returns sorted list with blank actual_count field for printing.
    """
    with get_conn() as conn:
        if project_id:
            # Get pick list for the project
            items = conn.execute("""
                SELECT pi.part_id, pi.qty AS project_qty, pi.item_type
                FROM project_items pi
                WHERE pi.project_id=? AND (pi.item_type IS NULL OR pi.item_type != 'DELETED')
            """, (project_id,)).fetchall()
            # Explode BOMs to get PRT/RAW leaf parts
            agg = defaultdict(float)
            for item in items:
                flat = explode_bom_flat(item[0], float(item[1] or 1))
                for row in flat:
                    if row['type'] in ('PRT','RAW') and not row['optional'] and row['total_qty'] > 0:
                        agg[row['part_id']] += row['total_qty']
            # Enrich
            result = []
            for part_id, need_qty in sorted(agg.items()):
                p = conn.execute("""
                    SELECT part_id, plain_desc, supplier, supplier_pn,
                           uom, unit_cost, qty_on_hand, category, type
                    FROM parts WHERE part_id=?
                """, (part_id,)).fetchone()
                if p:
                    result.append({**dict(p), 'need_qty': need_qty, 'actual_count': ''})
            return result
        else:
            q = """
                SELECT part_id, plain_desc, supplier, supplier_pn,
                       uom, unit_cost, qty_on_hand, category, type
                FROM parts
                WHERE type IN ('PRT','RAW')
                AND (status IS NULL OR status != 'OBSOLETE')
            """
            params = []
            if cat_filter:
                q += " AND category=?"
                params.append(cat_filter)
            q += " ORDER BY category, part_id"
            rows = conn.execute(q, params).fetchall()
            return [{**dict(r), 'need_qty': None, 'actual_count': ''} for r in rows]

def get_all_projects_with_summary() -> list[dict]:
    """Projects list enriched with financial summary.
    Quoted_total and gross_margin_pct are stored by the quote GET route on every
    visit — no calculation here, values always match what the quote page shows.
    Computed via a single DB connection rather than N+1 queries.
    """
    with get_conn() as conn:
        projects = conn.execute("SELECT * FROM projects ORDER BY created_at DESC").fetchall()
        projects = [dict(p) for p in projects]
        
        quotes_rows = conn.execute("SELECT project_id, quoted_total, gross_margin_pct, total_internal FROM project_quotes").fetchall()
        quotes_map = {q['project_id']: dict(q) for q in quotes_rows}
        
        items_rows = conn.execute("""
            SELECT pi.project_id, pi.qty, pi.item_type,
                   p.type, p.unit_cost, p.labor_hrs, p.rolled_labor_hrs
            FROM project_items pi
            JOIN parts p ON p.part_id = pi.part_id
        """).fetchall()
        
        other_items_rows = conn.execute("SELECT project_id, labor_hrs FROM project_other_items").fetchall()

    proj_totals = {}
    for r in items_rows:
        pid = r['project_id']
        if pid not in proj_totals:
            proj_totals[pid] = {'total_mat': 0.0, 'total_lbr': 0.0}
            
        qty = float(r['qty'])
        sign = -1.0 if r['item_type'] == 'DELETED' else 1.0
        eff_qty = qty * sign
        
        mat_cost = float(r['unit_cost'] or 0) * eff_qty
        if r['type'] in PART_TYPES_STATIC_COST:
            lbr_hrs = float(r['labor_hrs'] or 0) * eff_qty
        else:
            lbr_hrs = float(r['rolled_labor_hrs'] or 0) * eff_qty
            
        proj_totals[pid]['total_mat'] += mat_cost
        proj_totals[pid]['total_lbr'] += lbr_hrs
        
    for r in other_items_rows:
        pid = r['project_id']
        if pid not in proj_totals:
            proj_totals[pid] = {'total_mat': 0.0, 'total_lbr': 0.0}
        proj_totals[pid]['total_lbr'] += float(r['labor_hrs'] or 0)

    result = []
    for p in projects:
        pid = p['project_id']
        quote = quotes_map.get(pid, {})
        totals = proj_totals.get(pid, {'total_mat': 0.0, 'total_lbr': 0.0})
        
        mat = totals['total_mat']
        lbr_hrs = totals['total_lbr']
        lbr_cost = lbr_hrs * float(p.get('labor_rate') or 25)
        total_cost = mat + lbr_cost
        
        result.append({
            **p,
            'total_material':   mat,
            'total_labor_hrs':  lbr_hrs,
            'labor_cost':       lbr_cost,
            'total_cost':       total_cost,
            'total_internal':   float(quote.get('total_internal') or total_cost),
            'quoted_total':     float(quote.get('quoted_total') or 0),
            'quote_margin_pct': float(quote.get('gross_margin_pct') or 0),
        })
    return result

def get_project_summary(project_id: str) -> 'Optional[dict]':
    """Full cost/labor summary — now also computes per-line retail and discounts."""
    project = get_project(project_id)
    if not project: return None
    items = get_project_items(project_id)

    # This function is a data-gathering pass only.
    # All labor-rate-dependent math (labor cost, discounts, retail, margin)
    # is the exclusive responsibility of calculations.py, which has access
    # to quote.labor_rate_quoted -- the only rate the user ever sets.
    # project.labor_rate is a vestigial DB default ($25) never exposed in the UI;
    # doing pricing math with it here would produce wrong numbers.
    total_mat  = 0.0
    total_lbr  = 0.0
    line_items = []

    for item in items:
        bom      = calc_bom_summary(item['part_id'], item['qty'])
        mat_cost = bom['material_cost']
        lbr_hrs  = bom['labor_hrs']
        sign     = -1 if item['item_type'] == 'DELETED' else 1
        total_mat += sign * mat_cost
        total_lbr += sign * lbr_hrs
        line_items.append({**item,
            'material_cost': sign * mat_cost,
            'labor_hrs':     sign * lbr_hrs,
        })

    # total_labor_hrs folds in other-item hours so calculations.py
    # sees the full hour count when computing labor cost at the quoted rate.
    other_items   = get_project_other_items(project_id)
    other_lbr_hrs = sum(float(oi.get('labor_hrs') or 0) for oi in other_items)

    return {
        'project':         project,
        'line_items':      line_items,
        'total_material':  total_mat,
        'total_labor_hrs': total_lbr + other_lbr_hrs,
    }

def get_project_needs_map() -> dict[str, dict[str, float]]:
    with get_conn() as conn:
        active = conn.execute("SELECT project_id FROM projects WHERE status='ACTIVE'").fetchall()
    
    project_needs = {}
    for row in active:
        pid = row['project_id']
        pick = generate_pick_list(pid)
        needs = {}
        for p in pick:
            part_id = p['part_id']
            unpicked_qty = p['total_qty'] - p.get('picked_qty', 0.0)
            if p.get('picked') and p.get('picked_qty', 0.0) == 0.0:
                unpicked_qty = 0.0
            if unpicked_qty > 0:
                needs[part_id] = max(0.0, unpicked_qty)
        if needs:
            project_needs[pid] = needs
    return project_needs

def get_global_need() -> dict[str, float]:
    """
    Net qty of each PRT/RAW needed across all ACTIVE projects combined.
    DELETED items subtract from the total. Returns {part_id: net_qty >= 0}.
    """
    with get_conn() as conn:
        active = conn.execute(
            "SELECT project_id FROM projects WHERE status='ACTIVE'"
        ).fetchall()
    aggregated: dict[str, float] = {}

    for row in active:
        pick = generate_pick_list(row['project_id'])
        for p in pick:
            pid = p['part_id']
            # Exclude items that have already been picked, because their quantities
            # have already been deducted from qty_on_hand in inventory.
            unpicked_qty = p['total_qty'] - p.get('picked_qty', 0.0)
            
            # Fallback for legacy picks where picked=True but picked_qty=0
            if p.get('picked') and p.get('picked_qty', 0.0) == 0.0:
                unpicked_qty = 0.0
                
            if unpicked_qty > 0:
                aggregated[pid] = aggregated.get(pid, 0.0) + unpicked_qty

    return {k: max(0.0, v) for k, v in aggregated.items() if v > 0}

def get_part_field_values() -> dict[str, list[str]]:
    """Return unique non-empty values for base_desc, size_spec, variant — for autocomplete."""
    with get_conn() as conn:
        bd = [r[0] for r in conn.execute(
            "SELECT DISTINCT base_desc FROM parts WHERE base_desc != '' ORDER BY base_desc"
        ).fetchall()]
        ss = [r[0] for r in conn.execute(
            "SELECT DISTINCT size_spec FROM parts WHERE size_spec != '' ORDER BY size_spec"
        ).fetchall()]
        va = [r[0] for r in conn.execute(
            "SELECT DISTINCT variant FROM parts WHERE variant != '' ORDER BY variant"
        ).fetchall()]
    return {'base_desc': bd, 'size_spec': ss, 'variant': va}

def get_project_part_ids(project_id: str) -> set[str]:
    """Return the set of PRT/RAW part_ids with net_qty > 0 for a project."""
    pick = generate_pick_list(project_id)
    return {p['part_id'] for p in pick if p['total_qty'] > 0}


# ── Project Quotes ─────────────────────────────────────────────────────────────

def get_or_create_quote(project_id: str) -> dict:
    """Return the quote row for a project, creating a default one if absent."""
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM project_quotes WHERE project_id = ?", (project_id,)
        ).fetchone()
        if row:
            q = dict(row)
        else:
            conn.execute(
                "INSERT INTO project_quotes (project_id) VALUES (?)", (project_id,)
            )
            log_change(conn, 'quote', project_id, 'CREATE_QUOTE')
            row = conn.execute(
                "SELECT * FROM project_quotes WHERE project_id = ?", (project_id,)
            ).fetchone()
            q = dict(row)
    try:
        q['other_items'] = json.loads(q.get('other_items') or '[]')
    except Exception:
        q['other_items'] = []
    return q

def save_quote_totals(project_id: str, quoted_total: float, gross_margin_pct: float, total_internal: float = 0) -> None:
    """Lightweight update — writes the three computed display fields.
    Called on every quote GET so the projects list always reads correct values."""
    try:
        with get_conn() as conn:
            conn.execute(
                """UPDATE project_quotes
                   SET quoted_total=?, gross_margin_pct=?, total_internal=?
                   WHERE project_id=?""",
                (quoted_total, gross_margin_pct, total_internal, project_id)
            )
            log_change(conn, 'quote', project_id, 'SAVE_TOTALS', new_val=f"Total: {quoted_total:.2f}")
    except Exception as e:
        logging.warning("save_quote_totals failed for %s: %s", project_id, e)


def save_quote(project_id: str, data: dict) -> tuple[bool, str]:
    """Upsert quote data for a project."""
    QUOTE_LOGGED = ['overhead_rate','markup_pct','labor_rate_quoted',
                    'freight_inbound','freight_outbound','cal_gases_cost','cal_gases_freight',
                    'training_days','training_cost','training_notes',
                    'discount_pct','discount_flat','discount_note',
                    'status','currency','internal_notes']
    other = data.get('other_items', [])
    if isinstance(other, str):
        try: other = json.loads(other)
        except Exception: other = []
    try:
        with get_conn() as conn:
            # Optimistic Locking check
            if data.get('updated_at'):
                row = conn.execute("SELECT updated_at FROM project_quotes WHERE project_id=?", (project_id,)).fetchone()
                if row and row['updated_at'] != data['updated_at']:
                    return False, "Quote was modified by another user. Please refresh to see latest changes before saving."

            old = conn.execute("SELECT * FROM project_quotes WHERE project_id=?", (project_id,)).fetchone()
            old_row = dict(old) if old else {}

            conn.execute("""
                INSERT INTO project_quotes
                    (project_id,version,status,currency,overhead_rate,markup_pct,labor_rate_quoted,
                     freight_inbound,freight_outbound,
                     cal_gases_cost,cal_gases_freight,
                     training_days,training_cost,training_notes,
                     frozen_material,frozen_labor_hrs,costs_frozen,frozen_at,
                     other_items,internal_notes,proforma_header,proforma_footer,
                     discount_pct,discount_flat,discount_note,
                     total_internal,
                     quoted_total,gross_margin_pct,
                     updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(project_id) DO UPDATE SET
                    status=excluded.status,
                    currency=excluded.currency,
                    overhead_rate=excluded.overhead_rate,
                    markup_pct=excluded.markup_pct,
                    labor_rate_quoted=excluded.labor_rate_quoted,
                    freight_inbound=excluded.freight_inbound,
                    freight_outbound=excluded.freight_outbound,
                    cal_gases_cost=excluded.cal_gases_cost,
                    cal_gases_freight=excluded.cal_gases_freight,
                    training_days=excluded.training_days,
                    training_cost=excluded.training_cost,
                    training_notes=excluded.training_notes,
                    frozen_material=excluded.frozen_material,
                    frozen_labor_hrs=excluded.frozen_labor_hrs,
                    costs_frozen=excluded.costs_frozen,
                    frozen_at=excluded.frozen_at,
                    other_items=excluded.other_items,
                    internal_notes=excluded.internal_notes,
                    proforma_header=excluded.proforma_header,
                    proforma_footer=excluded.proforma_footer,
                    discount_pct=excluded.discount_pct,
                    discount_flat=excluded.discount_flat,
                    discount_note=excluded.discount_note,
                    total_internal=excluded.total_internal,
                    quoted_total=excluded.quoted_total,
                    gross_margin_pct=excluded.gross_margin_pct,
                    updated_at=excluded.updated_at
            """, (
                project_id,
                int(data.get('version') or 1),
                data.get('status', 'DRAFT'),
                data.get('currency', 'USD'),
                float(data.get('overhead_rate') or 1.0),
                float(data.get('markup_pct') if data.get('markup_pct') is not None else 0),
                float(data.get('labor_rate_quoted') or 0),
                float(data.get('freight_inbound') or 0),
                float(data.get('freight_outbound') or 0),
                float(data.get('cal_gases_cost') or 0),
                float(data.get('cal_gases_freight') or 0),
                float(data.get('training_days') or 0),
                float(data.get('training_cost') or 0),
                data.get('training_notes', ''),
                float(data.get('frozen_material') or 0),
                float(data.get('frozen_labor_hrs') or 0),
                int(bool(data.get('costs_frozen'))),
                data.get('frozen_at', ''),
                json.dumps(other),
                data.get('internal_notes', ''),
                data.get('proforma_header', ''),
                data.get('proforma_footer', ''),
                float(data.get('discount_pct') or 0),
                float(data.get('discount_flat') or 0),
                data.get('discount_note', ''),
                float(data.get('total_internal') or 0),
                float(data.get('quoted_total') or 0),
                float(data.get('gross_margin_pct') or 0),
                datetime.datetime.now().isoformat(timespec='seconds'),
            ))
            _diff_log(conn, 'quote', project_id, old_row, data, QUOTE_LOGGED)

            status = data.get('status', 'DRAFT')
            if status in ('SENT', 'ACCEPTED'):
                version = int(data.get('version') or 1)
                exists = conn.execute("SELECT 1 FROM quote_snapshots WHERE project_id=? AND version=?", (project_id, version)).fetchone()
                if not exists:
                    snap = build_print_project(project_id)
                    if snap:
                        conn.execute("INSERT INTO quote_snapshots (project_id, version, snapshot_json) VALUES (?, ?, ?)",
                            (project_id, version, json.dumps(snap))
                        )
        return True, "Quote saved."
    except Exception as e:
        return False, str(e)


def increment_quote_version(project_id: str) -> int:
    """Bump the quote version number (call when sending to customer)."""
    with get_conn() as conn:
        conn.execute(
            "UPDATE project_quotes SET version = version + 1 WHERE project_id = ?",
            (project_id,)
        )
        row = conn.execute(
            "SELECT version FROM project_quotes WHERE project_id = ?", (project_id,)
        ).fetchone()
        log_change(conn, 'quote', project_id, 'INCREMENT_VERSION', new_val=row['version'])
    return row['version'] if row else 1

# ── Optional Items Suggestion ─────────────────────────────────────────────────

def get_project_optional_items(project_id: str) -> list[dict]:
    """
    Returns a deduplicated list of optional BOM components (qty=0) that are
    reachable from any non-DELETED item currently in the project, but are NOT
    already present in the project's item list.

    Each result dict includes:
      part_id, plain_desc, type, unit_cost, uom, labor_hrs,
      source_parts  — list of parent part_ids that expose this optional
    """
    # Collect current project part_ids (to exclude already-added ones)
    current_items = get_project_items(project_id)
    existing_part_ids = {
        item['part_id'] for item in current_items
        if item['item_type'] != 'DELETED'
    }

    # Top-level assemblies/parts on the project that we'll walk
    top_level_ids = [
        item['part_id'] for item in current_items
        if item['item_type'] != 'DELETED'
    ]

    # Pre-load the full BOM adjacency list once — eliminates N+1 DB calls during recursion
    with get_conn() as conn:
        _all_bom = conn.execute("SELECT parent_id, child_id, qty FROM bom").fetchall()
    _adjacency: dict[str, list] = {}
    for _row in _all_bom:
        _adjacency.setdefault(_row['parent_id'], []).append((_row['child_id'], _row['qty']))

    def collect_optionals(parent_id: str, visited: set) -> list[dict]:
        if parent_id in visited:
            return []
        visited = visited | {parent_id}
        results = []
        for child_id, qty in _adjacency.get(parent_id, []):
            if qty == 0:
                # This is an optional item — record it
                results.append({'part_id': child_id, 'source_parent': parent_id})
            else:
                # Walk deeper through required children
                results.extend(collect_optionals(child_id, visited))
        return results

    # Walk all top-level project items
    found: dict[str, dict] = {}  # part_id -> {part_id, plain_desc, ..., source_parts:[...]}
    for top_id in top_level_ids:
        for opt in collect_optionals(top_id, set()):
            pid = opt['part_id']
            if pid in existing_part_ids:
                continue  # already in project — skip
            if pid not in found:
                found[pid] = {'part_id': pid, 'source_parts': []}
            src = opt['source_parent']
            if src not in found[pid]['source_parts']:
                found[pid]['source_parts'].append(src)

    if not found:
        return []

    # Enrich with part metadata in a single query
    ids = list(found.keys())
    placeholders = ','.join('?' * len(ids))
    with get_conn() as conn:
        rows = conn.execute(
            f"SELECT part_id, plain_desc, type, unit_cost, uom, labor_hrs"
            f" FROM parts WHERE part_id IN ({placeholders})",
            ids
        ).fetchall()

    enriched = []
    for row in rows:
        pid = row['part_id']
        entry = {**found[pid], **dict(row)}
        enriched.append(entry)

    # Sort by type then part_id for consistent ordering
    enriched.sort(key=lambda x: (x.get('type',''), x['part_id']))
    return enriched


# ── Packing Info ──────────────────────────────────────────────────────────────

def get_project_other_items(project_id: str) -> list[dict]:
    with get_conn() as conn:
        return [dict(r) for r in conn.execute(
            """SELECT * FROM project_other_items
               WHERE project_id=? ORDER BY sort_order, id""",
            (project_id,)
        ).fetchall()]

def add_project_other_item(project_id: str) -> tuple[bool, int]:
    try:
        with get_conn() as conn:
            cur = conn.execute(
                """INSERT INTO project_other_items (project_id, sort_order)
                   VALUES (?, COALESCE((SELECT MAX(sort_order)+1 FROM project_other_items WHERE project_id=?), 0))""",
                (project_id, project_id)
            )
            new_id = cur.lastrowid
            log_change(conn, 'project_other_item', project_id, 'create',
                       field='id', old_val=None, new_val=str(new_id))
            return True, new_id
    except Exception as e:
        return False, str(e)

def update_project_other_item(item_id: int, **kwargs) -> tuple[bool, str]:
    allowed = {'description','cost','labor_hrs','apply_markup','box_num',
               'discount_pct','discount_flat','show_on_proforma','sort_order'}
    fields = {k: v for k, v in kwargs.items() if k in allowed}
    if not fields:
        return False, "No valid fields"
    try:
        with get_conn() as conn:
            old = conn.execute("SELECT * FROM project_other_items WHERE id=?", (item_id,)).fetchone()
            old_row    = dict(old) if old else {}
            project_id = old_row.get('project_id', str(item_id))
            for field, value in fields.items():
                conn.execute(
                    f"UPDATE project_other_items SET {field}=? WHERE id=?",
                    (value, item_id)
                )
                log_change(conn, 'project_other_item', project_id, 'update',
                           field=field, old_val=old_row.get(field), new_val=value)
        return True, "Updated"
    except Exception as e:
        return False, str(e)

def delete_project_other_item(item_id: int) -> tuple[bool, str]:
    try:
        with get_conn() as conn:
            old = conn.execute("SELECT project_id, description FROM project_other_items WHERE id=?", (item_id,)).fetchone()
            conn.execute("DELETE FROM project_other_items WHERE id=?", (item_id,))
            if old:
                log_change(conn, 'project_other_item', old['project_id'], 'delete',
                           field='description', old_val=old['description'], new_val=None)
        return True, "Deleted"
    except Exception as e:
        return False, str(e)

def get_project_boxes(project_id: str) -> list[dict]:
    with get_conn() as conn:
        return [dict(r) for r in conn.execute(
            "SELECT * FROM project_boxes WHERE project_id=? ORDER BY box_num", (project_id,)
        ).fetchall()]

def get_project_pallets(project_id: str) -> list[dict]:
    with get_conn() as conn:
        return [dict(r) for r in conn.execute(
            "SELECT * FROM project_pallets WHERE project_id=? ORDER BY pallet_num", (project_id,)
        ).fetchall()]

def save_project_packing(project_id: str, boxes: list[dict], pallets: list[dict]) -> tuple[bool, str]:
    with get_conn() as conn:
        conn.execute("DELETE FROM project_boxes WHERE project_id=?", (project_id,))
        for b in boxes:
            conn.execute(
                "INSERT INTO project_boxes (project_id, box_num, weight, pallet_num) VALUES (?,?,?,?)",
                (project_id, b.get('box_num',''), float(b.get('weight') or 0), b.get('pallet_num',''))
            )
        conn.execute("DELETE FROM project_pallets WHERE project_id=?", (project_id,))
        for p in pallets:
            conn.execute(
                "INSERT INTO project_pallets (project_id, pallet_num, weight, dimensions) VALUES (?,?,?,?)",
                (project_id, p.get('pallet_num',''), float(p.get('weight') or 0), p.get('dimensions',''))
            )
        log_change(conn, 'project', project_id, 'SAVE_PACKING', new_val=f"{len(boxes)} boxes, {len(pallets)} pallets")
    return True, "Packing info saved."


def get_part_cost_history(part_id: str) -> list[dict]:
    with get_conn() as conn:
        return [dict(r) for r in conn.execute(
            "SELECT * FROM part_cost_history WHERE part_id=? ORDER BY changed_at DESC, id DESC", 
            (part_id,)
        ).fetchall()]

# ── Attachments ────────────────────────────────────────────────────────────────
def get_part_attachments(part_id: str) -> list[dict]:
    with get_conn() as conn:
        return [dict(r) for r in conn.execute(
            "SELECT * FROM part_attachments WHERE part_id=? ORDER BY uploaded_at DESC",
            (part_id,)
        ).fetchall()]

def get_project_attachments(project_id: str) -> list[dict]:
    # First, get all parts in the project BOM using explode_bom_flat
    with get_conn() as conn:
        bom_ctx = _build_bom_ctx(conn)
        items = conn.execute("SELECT part_id, qty FROM project_items WHERE project_id=? AND item_type != 'DELETED'", (project_id,)).fetchall()
        
        part_ids = set()
        for item in items:
            flat = explode_bom_flat(item['part_id'], qty_mult=1.0, _conn=conn, bom_ctx=bom_ctx)
            for r in flat:
                part_ids.add(r['part_id'])
        
        if not part_ids: return []
        
        placeholders = ','.join('?' for _ in part_ids)
        rows = conn.execute(
            f"SELECT * FROM part_attachments WHERE part_id IN ({placeholders}) ORDER BY uploaded_at DESC",
            tuple(part_ids)
        ).fetchall()
        return [dict(r) for r in rows]

def add_part_attachment(data: dict) -> bool:
    with get_conn() as conn:
        conn.execute('''
            INSERT INTO part_attachments (id, part_id, filename, original_filename, mime_type, size_bytes, uploaded_by)
            VALUES (:id, :part_id, :filename, :original_filename, :mime_type, :size_bytes, :uploaded_by)
        ''', data)
        log_change(conn, 'part', data['part_id'], 'ADD_ATTACHMENT', field=data['original_filename'])
        return True

def delete_part_attachment(att_id: str) -> dict | None:
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM part_attachments WHERE id=?", (att_id,)).fetchone()
        if not row: return None
        conn.execute("DELETE FROM part_attachments WHERE id=?", (att_id,))
        log_change(conn, 'part', row['part_id'], 'DELETE_ATTACHMENT', field=row['original_filename'])
        return dict(row)
